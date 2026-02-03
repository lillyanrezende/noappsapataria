import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from db import DB
from services import ProductService, AuthService, DomainService
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def safe_int(s, default=None):
    """Converte string para int com segurança"""
    try:
        return int(str(s).strip())
    except Exception:
        return default


def copy_to_clipboard(widget):
    """Copia texto selecionado para clipboard"""
    try:
        text = widget.get("sel.first", "sel.last")
        widget.clipboard_clear()
        widget.clipboard_append(text)
        widget.update()
        messagebox.showinfo("OK", "Texto copiado para clipboard!")
    except tk.TclError:
        messagebox.showwarning("Atenção", "Seleciona o texto que queres copiar.")


def paste_from_clipboard(widget):
    """Cola texto do clipboard"""
    try:
        text = widget.clipboard_get()
        widget.insert(tk.INSERT, text)
    except tk.TclError:
        messagebox.showwarning("Atenção", "Não há texto no clipboard.")


class AddOptionDialog(simpledialog.Dialog):
    """Dialog para adicionar nova opção a um domínio"""
    
    def __init__(self, parent, title, label):
        self.label = label
        self.value = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text=self.label).grid(row=0, column=0, sticky="w", padx=6, pady=6)
        self.entry = ttk.Entry(master, width=40)
        self.entry.grid(row=1, column=0, padx=6, pady=6)
        return self.entry

    def apply(self):
        self.value = self.entry.get().strip()


class App(tk.Tk):
    """Aplicação principal"""
    
    def __init__(self, db: DB):
        super().__init__()
        self.title("Sistema Sapataria - Cadastro/Stock (Supabase)")
        self.geometry("980x700")
        self.db = db
        self.user = None

        # Inicializar serviços
        self.product_service = ProductService(db)
        self.auth_service = AuthService(db)
        self.domain_service = DomainService(db)

        self.db.init_app_tables()
        self.show_login()

    def show_login(self):
        """Mostra tela de login"""
        for w in self.winfo_children():
            w.destroy()
        LoginFrame(self).pack(fill="both", expand=True)

    def show_main(self, user):
        """Mostra tela principal"""
        self.user = user
        for w in self.winfo_children():
            w.destroy()
        MainFrame(self).pack(fill="both", expand=True)


class LoginFrame(ttk.Frame):
    """Frame de login"""
    
    def __init__(self, app: App):
        super().__init__(app, padding=12)
        self.app = app

        ttk.Label(self, text="Login", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 12))

        ttk.Label(self, text="Username").grid(row=1, column=0, sticky="w")
        self.username = ttk.Entry(self, width=40)
        self.username.grid(row=2, column=0, sticky="w", pady=(0, 8))

        ttk.Label(self, text="Password").grid(row=3, column=0, sticky="w")
        self.password = ttk.Entry(self, width=40, show="*")
        self.password.grid(row=4, column=0, sticky="w", pady=(0, 12))

        btns = ttk.Frame(self)
        btns.grid(row=5, column=0, sticky="w")

        ttk.Button(btns, text="Entrar", command=self.do_login).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="Criar conta", command=self.do_register).grid(row=0, column=1)

        ttk.Separator(self).grid(row=6, column=0, sticky="ew", pady=18)

        info = (
            "Notas:\n"
            "- Utilizadores guardados em public.app_users\n"
            "- Senha guardada com hash (bcrypt)\n"
        )
        ttk.Label(self, text=info).grid(row=7, column=0, sticky="w")

    def do_login(self):
        """Realiza login"""
        u = self.username.get().strip()
        p = self.password.get().strip()
        if not u or not p:
            messagebox.showwarning("Atenção", "Preenche username e password.")
            return

        try:
            res = self.app.auth_service.authenticate(u, p)
            if res == "inactive":
                messagebox.showerror("Bloqueado", "Utilizador inativo. Contacta o administrador.")
                return
            if not res:
                messagebox.showerror("Erro", "Credenciais inválidas.")
                return

            self.app.db.audit(res, "LOGIN", "profiles", entity_pk=f"user_id={res['user_id']}", details={})
            self.app.show_main(res)
        except Exception as e:
            messagebox.showerror("Erro de Conexão", 
                               f"Não foi possível conectar ao Supabase.\n\n"
                               f"Verifica se o ficheiro .env tem as credenciais corretas:\n"
                               f"- SUPABASE_URL\n"
                               f"- SUPABASE_KEY\n\n"
                               f"Erro: {str(e)}")

    def do_register(self):
        """Regista novo utilizador"""
        username = simpledialog.askstring("Criar conta", "Username (login):", parent=self)
        if not username:
            return
        nome = simpledialog.askstring("Criar conta", "Nome do utilizador:", parent=self)
        if not nome:
            return
        setor = simpledialog.askstring("Criar conta", "Setor (opcional):", parent=self)
        password = simpledialog.askstring("Criar conta", "Password:", show="*", parent=self)
        if not password:
            return

        try:
            user_id = self.app.auth_service.create_user(username, nome, setor, password)
            self.app.db.audit({"user_id": user_id, "username": username}, "REGISTER", "profiles", 
                            entity_pk=f"user_id={user_id}", details={"username": username, "setor": setor})
            messagebox.showinfo("OK", "Conta criada. Agora faz login.")
        except Exception as e:
            messagebox.showerror("Erro de Conexão", 
                               f"Não foi possível criar o utilizador.\n\n"
                               f"Verifica se o ficheiro .env tem as credenciais corretas.\n\n"
                               f"Erro: {str(e)}")


class MainFrame(ttk.Frame):
    """Frame principal com tabs"""
    
    def __init__(self, app: App):
        super().__init__(app, padding=10)
        self.app = app

        top = ttk.Frame(self)
        top.pack(fill="x")

        ttk.Label(top, text=f"Utilizador: {app.user['nome_usuario']}  |  Setor: {app.user.get('setor') or '-'}", 
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        ttk.Button(top, text="Sair", command=self.logout).pack(side="right")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, pady=(10, 0))

        self.tab_create = CreateTab(nb, app)
        self.tab_update = UpdateTab(nb, app)
        self.tab_delete = DeleteTab(nb, app)
        self.tab_view = ViewTab(nb, app)
        self.tab_warehouse = WarehouseTab(nb, app)

        nb.add(self.tab_create, text="Cadastrar")
        nb.add(self.tab_update, text="Alterar")
        nb.add(self.tab_delete, text="Excluir")
        nb.add(self.tab_view, text="Visualizar")
        nb.add(self.tab_warehouse, text="Armazéns")

    def logout(self):
        """Realiza logout"""
        self.app.db.audit(self.app.user, "LOGOUT", "profiles", 
                         entity_pk=f"user_id={self.app.user['user_id']}", details={})
        self.app.user = None
        self.app.show_login()


class BaseTab(ttk.Frame):
    """Tab base com funcionalidades comuns"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, padding=10)
        self.app = app
        self.db = app.db
        self.domain_service = app.domain_service
        self.cache = {}

    def load_domains(self):
        """Carrega todos os domínios"""
        self.cache["brands"] = self.domain_service.get_domain_list("brands")
        self.cache["categories"] = self.domain_service.get_domain_list("categories")
        self.cache["colors"] = self.domain_service.get_domain_list("colors")
        self.cache["sizes"] = self.domain_service.get_domain_list("sizes")
        self.cache["warehouses"] = self.domain_service.get_domain_list("warehouses")
        self.cache["suppliers"] = self.domain_service.get_domain_list("suppliers")

    def tuple_list_to_map(self, rows):
        """Converte lista de tuples em mapas id->name e name->id"""
        id_to_name = {r[0]: r[1] for r in rows}
        name_to_id = {r[1]: r[0] for r in rows}
        return id_to_name, name_to_id

    def ask_add_option(self, table, label):
        """Pergunta ao utilizador para adicionar nova opção"""
        dlg = AddOptionDialog(self, f"Adicionar {label}", f"Novo {label}:")
        if dlg.value:
            if table == "subcategories":
                raise RuntimeError("Use add_subcategory (depende da categoria).")
            new_id = self.domain_service.add_domain_value(table, dlg.value)
            return new_id, dlg.value
        return None, None


class CreateTab(BaseTab):
    """Tab para criar produtos"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, app)

        self.load_domains()

        ttk.Label(self, text="Cadastrar Produto", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12))

        # Fields
        self.gtin = ttk.Entry(self, width=28)
        self.ref_keyinvoice = ttk.Entry(self, width=28)
        self.nome_modelo = ttk.Entry(self, width=28)
        self.stock = ttk.Entry(self, width=10)

        # Dropdown variables
        self.var_brand = tk.StringVar()
        self.var_category = tk.StringVar()
        self.var_subcategory = tk.StringVar()
        self.var_color = tk.StringVar()
        self.var_size = tk.StringVar()
        self.var_warehouse = tk.StringVar()
        self.var_supplier = tk.StringVar()

        self._build_form()
        self._refresh_dropdowns()

    def _build_form(self):
        """Constrói formulário"""
        r = 1

        ttk.Label(self, text="GTIN").grid(row=r, column=0, sticky="w")
        self.gtin.grid(row=r, column=1, sticky="w", padx=(0, 12))
        ttk.Label(self, text="Ref KeyInvoice (opcional)").grid(row=r, column=2, sticky="w")
        self.ref_keyinvoice.grid(row=r, column=3, sticky="w")
        r += 1

        ttk.Label(self, text="Modelo (nome_modelo)").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.nome_modelo.grid(row=r, column=1, sticky="w", padx=(0, 12), pady=(8, 0))
        r += 1

        # Brand / Supplier
        ttk.Label(self, text="Marca").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.combo_brand = ttk.Combobox(self, textvariable=self.var_brand, width=25, state="readonly")
        self.combo_brand.grid(row=r, column=1, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_brand).grid(row=r, column=1, sticky="e", pady=(8, 0))

        ttk.Label(self, text="Fornecedor").grid(row=r, column=2, sticky="w", pady=(8, 0))
        self.combo_supplier = ttk.Combobox(self, textvariable=self.var_supplier, width=25, state="readonly")
        self.combo_supplier.grid(row=r, column=3, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_supplier).grid(row=r, column=3, sticky="e", pady=(8, 0))
        r += 1

        # Category / Subcategory
        ttk.Label(self, text="Categoria").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.combo_category = ttk.Combobox(self, textvariable=self.var_category, width=25, state="readonly")
        self.combo_category.grid(row=r, column=1, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_category).grid(row=r, column=1, sticky="e", pady=(8, 0))

        ttk.Label(self, text="Subcategoria").grid(row=r, column=2, sticky="w", pady=(8, 0))
        self.combo_subcategory = ttk.Combobox(self, textvariable=self.var_subcategory, width=25, state="readonly")
        self.combo_subcategory.grid(row=r, column=3, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_subcategory).grid(row=r, column=3, sticky="e", pady=(8, 0))
        r += 1

        # Color / Size / Warehouse / Stock
        ttk.Label(self, text="Cor").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.combo_color = ttk.Combobox(self, textvariable=self.var_color, width=25, state="readonly")
        self.combo_color.grid(row=r, column=1, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_color).grid(row=r, column=1, sticky="e", pady=(8, 0))

        ttk.Label(self, text="Tamanho").grid(row=r, column=2, sticky="w", pady=(8, 0))
        self.combo_size = ttk.Combobox(self, textvariable=self.var_size, width=25, state="readonly")
        self.combo_size.grid(row=r, column=3, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_size).grid(row=r, column=3, sticky="e", pady=(8, 0))
        r += 1

        ttk.Label(self, text="Armazém").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.combo_warehouse = ttk.Combobox(self, textvariable=self.var_warehouse, width=25, state="readonly")
        self.combo_warehouse.grid(row=r, column=1, sticky="w", padx=(0, 6), pady=(8, 0))
        ttk.Button(self, text="+", width=3, command=self.add_warehouse).grid(row=r, column=1, sticky="e", pady=(8, 0))

        ttk.Label(self, text="Stock").grid(row=r, column=2, sticky="w", pady=(8, 0))
        self.stock.grid(row=r, column=3, sticky="w", pady=(8, 0))
        r += 1

        ttk.Separator(self).grid(row=r, column=0, columnspan=6, sticky="ew", pady=14)
        r += 1

        ttk.Button(self, text="Cadastrar / Atualizar Stock", command=self.save).grid(row=r, column=0, sticky="w")
        ttk.Button(self, text="Limpar", command=self.clear).grid(row=r, column=1, sticky="w", padx=(8, 0))

        # events
        self.combo_category.bind("<<ComboboxSelected>>", lambda e: self._refresh_subcategories())

    def _refresh_dropdowns(self):
        """Atualiza dropdowns"""
        self.load_domains()

        self.brand_rows = self.cache["brands"]
        self.category_rows = self.cache["categories"]
        self.color_rows = self.cache["colors"]
        self.size_rows = self.cache["sizes"]
        self.warehouse_rows = self.cache["warehouses"]
        self.supplier_rows = self.cache["suppliers"]

        self.brand_id_to_name, self.brand_name_to_id = self.tuple_list_to_map(self.brand_rows)
        self.category_id_to_name, self.category_name_to_id = self.tuple_list_to_map(self.category_rows)
        self.color_id_to_name, self.color_name_to_id = self.tuple_list_to_map(self.color_rows)
        self.size_id_to_name, self.size_name_to_id = self.tuple_list_to_map(self.size_rows)
        self.warehouse_id_to_name, self.warehouse_name_to_id = self.tuple_list_to_map(self.warehouse_rows)
        self.supplier_id_to_name, self.supplier_name_to_id = self.tuple_list_to_map(self.supplier_rows)

        self.combo_brand["values"] = list(self.brand_name_to_id.keys())
        self.combo_category["values"] = list(self.category_name_to_id.keys())
        self.combo_color["values"] = list(self.color_name_to_id.keys())
        self.combo_size["values"] = list(self.size_name_to_id.keys())
        self.combo_warehouse["values"] = list(self.warehouse_name_to_id.keys())
        self.combo_supplier["values"] = list(self.supplier_name_to_id.keys())

        # default selections
        if not self.var_brand.get() and self.combo_brand["values"]:
            self.var_brand.set(self.combo_brand["values"][0])
        if not self.var_category.get() and self.combo_category["values"]:
            self.var_category.set(self.combo_category["values"][0])
            self._refresh_subcategories()
        if not self.var_color.get() and self.combo_color["values"]:
            self.var_color.set(self.combo_color["values"][0])
        if not self.var_size.get() and self.combo_size["values"]:
            self.var_size.set(self.combo_size["values"][0])
        if not self.var_warehouse.get() and self.combo_warehouse["values"]:
            self.var_warehouse.set(self.combo_warehouse["values"][0])
        if not self.var_supplier.get() and self.combo_supplier["values"]:
            self.var_supplier.set(self.combo_supplier["values"][0])

    def _refresh_subcategories(self):
        """Atualiza subcategorias"""
        cat_name = self.var_category.get().strip()
        cat_id = self.category_name_to_id.get(cat_name)
        if not cat_id:
            self.combo_subcategory["values"] = []
            self.var_subcategory.set("")
            return
        subs = self.domain_service.get_subcategories_by_category(cat_id)
        self.sub_id_to_name = {r[0]: r[1] for r in subs}
        self.sub_name_to_id = {r[1]: r[0] for r in subs}
        self.combo_subcategory["values"] = list(self.sub_name_to_id.keys())
        if self.combo_subcategory["values"]:
            self.var_subcategory.set(self.combo_subcategory["values"][0])
        else:
            self.var_subcategory.set("")

    # Add domain handlers
    def add_brand(self):
        new_id, new_val = self.ask_add_option("brands", "Marca")
        if new_id:
            self._refresh_dropdowns()
            self.var_brand.set(new_val)

    def add_supplier(self):
        new_id, new_val = self.ask_add_option("suppliers", "Fornecedor")
        if new_id:
            self._refresh_dropdowns()
            self.var_supplier.set(new_val)

    def add_category(self):
        new_id, new_val = self.ask_add_option("categories", "Categoria")
        if new_id:
            self._refresh_dropdowns()
            self.var_category.set(new_val)
            self._refresh_subcategories()

    def add_subcategory(self):
        cat_name = self.var_category.get().strip()
        cat_id = self.category_name_to_id.get(cat_name)
        if not cat_id:
            messagebox.showwarning("Atenção", "Escolhe uma categoria primeiro.")
            return
        dlg = AddOptionDialog(self, "Adicionar Subcategoria", "Nova Subcategoria:")
        if dlg.value:
            sub_id = self.domain_service.add_subcategory(cat_id, dlg.value)
            self._refresh_subcategories()
            self.var_subcategory.set(dlg.value)

    def add_color(self):
        new_id, new_val = self.ask_add_option("colors", "Cor")
        if new_id:
            self._refresh_dropdowns()
            self.var_color.set(new_val)

    def add_size(self):
        dlg = AddOptionDialog(self, "Adicionar Tamanho", "Novo Tamanho (ex: 40):")
        if dlg.value:
            new_id = self.domain_service.add_domain_value("sizes", dlg.value)
            self._refresh_dropdowns()
            self.var_size.set(dlg.value)

    def add_warehouse(self):
        new_id, new_val = self.ask_add_option("warehouses", "Armazém")
        if new_id:
            self._refresh_dropdowns()
            self.var_warehouse.set(new_val)

    def clear(self):
        """Limpa formulário"""
        self.gtin.delete(0, tk.END)
        self.ref_keyinvoice.delete(0, tk.END)
        self.nome_modelo.delete(0, tk.END)
        self.stock.delete(0, tk.END)

    def save(self):
        """Grava produto"""
        gtin = self.gtin.get().strip()
        if not gtin:
            messagebox.showwarning("Atenção", "GTIN é obrigatório.")
            return

        stock_val = safe_int(self.stock.get(), None)
        if stock_val is None or stock_val < 0:
            messagebox.showwarning("Atenção", "Stock inválido (precisa ser número inteiro >= 0).")
            return

        brand_id = self.brand_name_to_id.get(self.var_brand.get())
        cat_id = self.category_name_to_id.get(self.var_category.get())
        sub_id = self.sub_name_to_id.get(self.var_subcategory.get()) if hasattr(self, "sub_name_to_id") else None
        color_id = self.color_name_to_id.get(self.var_color.get())
        size_id = self.size_name_to_id.get(self.var_size.get())
        wh_id = self.warehouse_name_to_id.get(self.var_warehouse.get())
        supplier_id = self.supplier_name_to_id.get(self.var_supplier.get())

        nome_modelo = self.nome_modelo.get().strip()
        if not nome_modelo:
            messagebox.showwarning("Atenção", "Modelo (nome_modelo) é obrigatório.")
            return
        if not (brand_id and cat_id and sub_id and color_id and size_id and wh_id and supplier_id):
            messagebox.showwarning("Atenção", "Preenche todas as opções (marca/categoria/subcategoria/cor/tamanho/armazém/fornecedor).")
            return

        ref_keyinvoice = self.ref_keyinvoice.get().strip() or None

        try:
            success, message, variant_id = self.app.product_service.create_or_update_product(
                gtin=gtin,
                nome_modelo=nome_modelo,
                brand_id=brand_id,
                category_id=cat_id,
                subcategory_id=sub_id,
                supplier_id=supplier_id,
                color_id=color_id,
                size_id=size_id,
                warehouse_id=wh_id,
                stock=stock_val,
                ref_keyinvoice=ref_keyinvoice
            )

            if success:
                self.app.db.audit(self.app.user, "CREATE_OR_UPDATE_PRODUCT", "product_variant",
                                entity_pk=f"id={variant_id}",
                                details={"gtin": gtin, "stock": stock_val, "warehouse": self.var_warehouse.get()})
                messagebox.showinfo("OK", message)
            else:
                messagebox.showerror("Erro", message)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao cadastrar.\n\n{e}")


class UpdateTab(BaseTab):
    """Tab para atualizar produtos"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, app)

        self.load_domains()

        ttk.Label(self, text="Alterar Produto", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12)
        )

        ttk.Label(self, text="Buscar por:").grid(row=1, column=0, sticky="w")

        self.var_search_type = tk.StringVar(value="gtin")
        search_frame = ttk.Frame(self)
        search_frame.grid(row=1, column=1, columnspan=3, sticky="w")

        ttk.Radiobutton(search_frame, text="GTIN", variable=self.var_search_type, 
                       value="gtin").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type, 
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type, 
                       value="ref_woocommerce").pack(side="left")

        ttk.Label(self, text="Código:").grid(row=2, column=0, sticky="w", pady=(6, 0))
        self.search_value = ttk.Entry(self, width=28)
        self.search_value.grid(row=2, column=1, sticky="w", padx=(0, 8), pady=(6, 0))

        ttk.Button(self, text="Procurar", command=self.search_variants).grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Button(self, text="Carregar selecionado", command=self.load_selected_variant).grid(
            row=2, column=3, sticky="w", pady=(6, 0))
        ttk.Button(self, text="Alteração em Massa", command=self.bulk_update_stock).grid(
            row=2, column=4, sticky="w", padx=(8, 0), pady=(6, 0))

        # Treeview
        cols = ("variant_id", "gtin", "modelo", "cor", "tamanho", "ref_keyinvoice", "ref_woocomerce")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=6)
        self.tree.grid(row=3, column=0, columnspan=6, sticky="ew", pady=(10, 0))

        self.tree.heading("variant_id", text="Variant ID")
        self.tree.heading("gtin", text="GTIN")
        self.tree.heading("modelo", text="Modelo")
        self.tree.heading("cor", text="Cor")
        self.tree.heading("tamanho", text="Tamanho")
        self.tree.heading("ref_keyinvoice", text="Ref KeyInvoice")
        self.tree.heading("ref_woocomerce", text="Ref WooCommerce")

        self.tree.column("variant_id", width=90, anchor="w")
        self.tree.column("gtin", width=120, anchor="w")
        self.tree.column("modelo", width=180, anchor="w")
        self.tree.column("cor", width=120, anchor="w")
        self.tree.column("tamanho", width=80, anchor="w")
        self.tree.column("ref_keyinvoice", width=120, anchor="w")
        self.tree.column("ref_woocomerce", width=130, anchor="w")

        self.tree.bind("<Double-1>", lambda e: self.load_selected_variant())

        ttk.Separator(self).grid(row=4, column=0, columnspan=6, sticky="ew", pady=12)

        self.lbl_gtin = ttk.Label(self, text="Item carregado: -", font=("Segoe UI", 10, "bold"))
        self.lbl_gtin.grid(row=5, column=0, columnspan=6, sticky="w")
        self.form_row0 = 6

        # Form fields
        self.nome_modelo = ttk.Entry(self, width=28)
        self.ref_keyinvoice = ttk.Entry(self, width=28)

        self.var_brand = tk.StringVar()
        self.var_category = tk.StringVar()
        self.var_subcategory = tk.StringVar()
        self.var_supplier = tk.StringVar()
        self.var_color = tk.StringVar()
        self.var_size = tk.StringVar()
        self.var_warehouse = tk.StringVar()

        self.combo_brand = ttk.Combobox(self, textvariable=self.var_brand, width=25, state="readonly")
        self.combo_category = ttk.Combobox(self, textvariable=self.var_category, width=25, state="readonly")
        self.combo_subcategory = ttk.Combobox(self, textvariable=self.var_subcategory, width=25, state="readonly")
        self.combo_supplier = ttk.Combobox(self, textvariable=self.var_supplier, width=25, state="readonly")
        self.combo_color = ttk.Combobox(self, textvariable=self.var_color, width=25, state="readonly")
        self.combo_size = ttk.Combobox(self, textvariable=self.var_size, width=25, state="readonly")
        self.combo_warehouse = ttk.Combobox(self, textvariable=self.var_warehouse, width=25, state="readonly")

        self._refresh_dropdowns()
        self._build_edit_fields()

        self.combo_category.bind("<<ComboboxSelected>>", lambda e: self._refresh_subcategories())

        # Stock section
        r0 = self.form_row0
        ttk.Label(self, text="Gestão de Stock por Armazém", font=("Segoe UI", 10, "bold")).grid(
            row=r0+8, column=0, columnspan=6, sticky="w", pady=(14, 6))
        
        ttk.Label(self, text="Armazém:").grid(row=r0+9, column=0, sticky="w")
        self.combo_warehouse.grid(row=r0+9, column=1, sticky="w")
        
        ttk.Label(self, text="Quantidade:").grid(row=r0+9, column=2, sticky="w", padx=(12, 0))
        self.stock = ttk.Entry(self, width=12)
        self.stock.grid(row=r0+9, column=3, sticky="w")
        
        ttk.Button(self, text="➕ Adicionar", command=self.add_to_stock).grid(row=r0+9, column=4, padx=(8, 0))
        ttk.Button(self, text="➖ Retirar", command=self.remove_from_stock).grid(row=r0+9, column=5, padx=(4, 0))
        
        ttk.Button(self, text="📊 Ver Stock", command=self.show_current_stock).grid(row=r0+10, column=0, sticky="w", pady=(8, 0))
        
        # Área de visualização de stock
        ttk.Label(self, text="Stock nos Armazéns:", font=("Segoe UI", 9, "bold")).grid(
            row=r0+11, column=0, columnspan=6, sticky="w", pady=(12, 4))
        
        self.txt_stock_info = tk.Text(self, height=6, width=80, state="disabled")
        self.txt_stock_info.grid(row=r0+12, column=0, columnspan=6, sticky="ew", pady=(0, 8))

        self.loaded_variant_id = None
        self.loaded_gtin = None

    def search_variants(self):
        """Pesquisa variantes"""
        value = self.search_value.get().strip()
        search_type = self.var_search_type.get()

        if not value:
            messagebox.showwarning("Atenção", "Informa um código para buscar.")
            return

        try:
            results = self.db.search_variants(value, search_type)

            for iid in self.tree.get_children():
                self.tree.delete(iid)

            if not results:
                messagebox.showinfo("Não encontrado", "Nenhuma variação encontrada para esse código.")
                return

            for r in results:
                self.tree.insert(
                    "", "end",
                    values=(
                        r["variant_id"],
                        r.get("gtin") or "",
                        r.get("nome_modelo") or "",
                        r.get("cor") or "",
                        r.get("tamanho") or "",
                        r.get("ref_keyinvoice") or "",
                        r.get("ref_woocomerce") or "",
                    )
                )

            if len(results) == 1:
                self.tree.selection_set(self.tree.get_children()[0])
                self.load_selected_variant()

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao buscar.\n\n{e}")

    def load_selected_variant(self):
        """Carrega variante selecionada"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Seleciona uma variação na lista.")
            return

        values = self.tree.item(sel[0], "values")
        variant_id = values[0]

        try:
            full = self.db.get_full_view_by_variant_id(variant_id)
            if not full:
                messagebox.showinfo("Não encontrado", "Essa variação não existe mais.")
                return

            header, stocks = full
            self.loaded_variant_id = header["variant_id"]
            self.loaded_gtin = header.get("gtin")

            self.lbl_gtin.config(
                text=f"Item carregado: GTIN={header.get('gtin')} | variant_id={header['variant_id']}"
            )

            self.nome_modelo.delete(0, tk.END)
            self.nome_modelo.insert(0, header.get("nome_modelo") or "")

            self.ref_keyinvoice.delete(0, tk.END)
            self.ref_keyinvoice.insert(0, header.get("ref_keyinvoice") or "")

            self.var_brand.set(self.brand_id_to_name.get(header["marca_id"], ""))
            self.var_category.set(self.category_id_to_name.get(header["categoria_id"], ""))

            self._refresh_subcategories()
            self.var_subcategory.set(self.sub_id_to_name.get(header["subcategoria_id"], ""))

            self.var_supplier.set(self.supplier_id_to_name.get(header["fornecedor_id"], ""))
            self.var_color.set(self.color_id_to_name.get(header["cor_id"], ""))
            self.var_size.set(self.size_id_to_name.get(header["tamanho_id"], ""))

            # Atualizar área de stock
            self.txt_stock_info.config(state="normal")
            self.txt_stock_info.delete("1.0", tk.END)
            
            if stocks:
                self.txt_stock_info.insert(tk.END, "Stock disponível nos armazéns:\n\n")
                total_stock = 0
                for s in stocks:
                    self.txt_stock_info.insert(tk.END, f"  • {s['armazem']}: {s['stock']} unidades\n")
                    total_stock += s['stock']
                self.txt_stock_info.insert(tk.END, f"\n📦 Total geral: {total_stock} unidades")
                
                # Selecionar primeiro armazém com stock
                wh_name = stocks[0]["armazem"]
                if wh_name in self.combo_warehouse["values"]:
                    self.var_warehouse.set(wh_name)
                self.stock.delete(0, tk.END)
                self.stock.insert(0, str(stocks[0]["stock"]))
            else:
                self.txt_stock_info.insert(tk.END, "⚠️ Nenhum stock registado em armazéns.\n\n")
                self.txt_stock_info.insert(tk.END, "Use os botões 'Adicionar' para criar stock.")
                self.stock.delete(0, tk.END)
            
            self.txt_stock_info.config(state="disabled")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar.\n\n{e}")

    def _refresh_dropdowns(self):
        """Atualiza dropdowns"""
        self.load_domains()

        self.brand_rows = self.cache["brands"]
        self.category_rows = self.cache["categories"]
        self.color_rows = self.cache["colors"]
        self.size_rows = self.cache["sizes"]
        self.warehouse_rows = self.cache["warehouses"]
        self.supplier_rows = self.cache["suppliers"]

        self.brand_id_to_name, self.brand_name_to_id = self.tuple_list_to_map(self.brand_rows)
        self.category_id_to_name, self.category_name_to_id = self.tuple_list_to_map(self.category_rows)
        self.color_id_to_name, self.color_name_to_id = self.tuple_list_to_map(self.color_rows)
        self.size_id_to_name, self.size_name_to_id = self.tuple_list_to_map(self.size_rows)
        self.warehouse_id_to_name, self.warehouse_name_to_id = self.tuple_list_to_map(self.warehouse_rows)
        self.supplier_id_to_name, self.supplier_name_to_id = self.tuple_list_to_map(self.supplier_rows)

        self.combo_brand["values"] = list(self.brand_name_to_id.keys())
        self.combo_category["values"] = list(self.category_name_to_id.keys())
        self.combo_color["values"] = list(self.color_name_to_id.keys())
        self.combo_size["values"] = list(self.size_name_to_id.keys())
        self.combo_warehouse["values"] = list(self.warehouse_name_to_id.keys())
        self.combo_supplier["values"] = list(self.supplier_name_to_id.keys())

        if self.combo_warehouse["values"] and not self.var_warehouse.get():
            self.var_warehouse.set(self.combo_warehouse["values"][0])

    def _refresh_subcategories(self):
        """Atualiza subcategorias"""
        cat_name = self.var_category.get().strip()
        cat_id = self.category_name_to_id.get(cat_name)
        if not cat_id:
            self.combo_subcategory["values"] = []
            self.var_subcategory.set("")
            return
        subs = self.domain_service.get_subcategories_by_category(cat_id)
        self.sub_id_to_name = {r[0]: r[1] for r in subs}
        self.sub_name_to_id = {r[1]: r[0] for r in subs}
        self.combo_subcategory["values"] = list(self.sub_name_to_id.keys())
        if self.combo_subcategory["values"]:
            self.var_subcategory.set(self.combo_subcategory["values"][0])
        else:
            self.var_subcategory.set("")

    def _build_edit_fields(self):
        """Constrói campos de edição"""
        r0 = self.form_row0
        
        ttk.Label(self, text="Modelo").grid(row=r0+0, column=0, sticky="w", pady=(8, 0))
        self.nome_modelo.grid(row=r0+0, column=1, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Marca").grid(row=r0+1, column=0, sticky="w", pady=(8, 0))
        self.combo_brand.grid(row=r0+1, column=1, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Fornecedor").grid(row=r0+2, column=0, sticky="w", pady=(8, 0))
        self.combo_supplier.grid(row=r0+2, column=1, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Categoria").grid(row=r0+3, column=0, sticky="w", pady=(8, 0))
        self.combo_category.grid(row=r0+3, column=1, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Subcategoria").grid(row=r0+4, column=0, sticky="w", pady=(8, 0))
        self.combo_subcategory.grid(row=r0+4, column=1, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Cor").grid(row=r0+0, column=2, sticky="w", pady=(8, 0))
        self.combo_color.grid(row=r0+0, column=3, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Tamanho").grid(row=r0+1, column=2, sticky="w", pady=(8, 0))
        self.combo_size.grid(row=r0+1, column=3, sticky="w", pady=(8, 0))

        ttk.Label(self, text="Ref KeyInvoice").grid(row=r0+2, column=2, sticky="w", pady=(8, 0))
        self.ref_keyinvoice.grid(row=r0+2, column=3, sticky="w", pady=(8, 0))

        ttk.Separator(self).grid(row=r0+5, column=0, columnspan=6, sticky="ew", pady=12)

        ttk.Button(self, text="Guardar alterações (item)", command=self.save_item).grid(
            row=r0+6, column=0, sticky="w", pady=(14, 0))
        ttk.Button(self, text="Limpar", command=self.clear).grid(
            row=r0+6, column=1, sticky="w", padx=(8, 0), pady=(14, 0))

    def clear(self):
        """Limpa formulário"""
        self.loaded_variant_id = None
        self.loaded_gtin = None
        self.lbl_gtin.config(text="Item carregado: -")
        self.nome_modelo.delete(0, tk.END)
        self.ref_keyinvoice.delete(0, tk.END)
        self.stock.delete(0, tk.END)
        
        # Limpar área de stock
        self.txt_stock_info.config(state="normal")
        self.txt_stock_info.delete("1.0", tk.END)
        self.txt_stock_info.config(state="disabled")

    def save_item(self):
        """Grava alterações do item"""
        if not self.loaded_variant_id or not self.loaded_gtin:
            messagebox.showwarning("Atenção", "Carrega um produto primeiro.")
            return

        gtin = self.loaded_gtin
        nome_modelo = self.nome_modelo.get().strip()
        if not nome_modelo:
            messagebox.showwarning("Atenção", "Modelo é obrigatório.")
            return

        marca_id = self.brand_name_to_id.get(self.var_brand.get())
        categoria_id = self.category_name_to_id.get(self.var_category.get())
        subcategoria_id = self.sub_name_to_id.get(self.var_subcategory.get()) if hasattr(self, "sub_name_to_id") else None
        fornecedor_id = self.supplier_name_to_id.get(self.var_supplier.get())

        cor_id = self.color_name_to_id.get(self.var_color.get())
        tamanho_id = self.size_name_to_id.get(self.var_size.get())
        ref_keyinvoice = self.ref_keyinvoice.get().strip() or None

        if not all([marca_id, categoria_id, subcategoria_id, fornecedor_id, cor_id, tamanho_id]):
            messagebox.showwarning("Atenção", "Preenche marca/categoria/subcategoria/fornecedor/cor/tamanho.")
            return

        ok, msg = self.app.product_service.update_product_details(
            gtin=gtin,
            model_fields={
                "nome_modelo": nome_modelo,
                "marca_id": marca_id,
                "categoria_id": categoria_id,
                "subcategoria_id": subcategoria_id,
                "fornecedor_id": fornecedor_id
            },
            variant_fields={
                "cor_id": cor_id,
                "tamanho_id": tamanho_id,
                "ref_keyinvoice": ref_keyinvoice
            }
        )

        if ok:
            self.app.db.audit(self.app.user, "UPDATE_ITEM", "product_variant",
                            entity_pk=f"gtin={gtin}",
                            details={"fields": ["model", "variant"]})
            messagebox.showinfo("OK", msg)
        else:
            messagebox.showerror("Erro", msg)

    def add_to_stock(self):
        """Adiciona quantidade ao stock"""
        if not self.loaded_variant_id:
            messagebox.showwarning("Atenção", "Carrega um produto primeiro.")
            return

        quantity = safe_int(self.stock.get(), None)
        if quantity is None or quantity <= 0:
            messagebox.showwarning("Atenção", "Quantidade inválida. Deve ser maior que 0.")
            return

        wh_id = self.warehouse_name_to_id.get(self.var_warehouse.get())
        if not wh_id:
            messagebox.showwarning("Atenção", "Escolhe um armazém.")
            return

        try:
            success, msg = self.app.product_service.add_to_stock(self.loaded_variant_id, wh_id, quantity)
            if success:
                self.app.db.audit(self.app.user, "ADD_STOCK", "warehouse_stock",
                                entity_pk=f"variant_id={self.loaded_variant_id},warehouse_id={wh_id}",
                                details={"quantity_added": quantity})
                messagebox.showinfo("OK", msg)
                self.stock.delete(0, tk.END)
                # Atualizar visualização de stock
                self._refresh_stock_display()
            else:
                messagebox.showerror("Erro", msg)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao adicionar stock.\n\n{e}")

    def remove_from_stock(self):
        """Remove quantidade do stock"""
        if not self.loaded_variant_id:
            messagebox.showwarning("Atenção", "Carrega um produto primeiro.")
            return

        quantity = safe_int(self.stock.get(), None)
        if quantity is None or quantity <= 0:
            messagebox.showwarning("Atenção", "Quantidade inválida. Deve ser maior que 0.")
            return

        wh_id = self.warehouse_name_to_id.get(self.var_warehouse.get())
        if not wh_id:
            messagebox.showwarning("Atenção", "Escolhe um armazém.")
            return

        try:
            success, msg = self.app.product_service.remove_from_stock(self.loaded_variant_id, wh_id, quantity)
            if success:
                self.app.db.audit(self.app.user, "REMOVE_STOCK", "warehouse_stock",
                                entity_pk=f"variant_id={self.loaded_variant_id},warehouse_id={wh_id}",
                                details={"quantity_removed": quantity})
                messagebox.showinfo("OK", msg)
                self.stock.delete(0, tk.END)
                # Atualizar visualização de stock
                self._refresh_stock_display()
            else:
                messagebox.showerror("Erro", msg)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao retirar stock.\n\n{e}")

    def show_current_stock(self):
        """Mostra stock atual do armazém selecionado"""
        if not self.loaded_variant_id:
            messagebox.showwarning("Atenção", "Carrega um produto primeiro.")
            return

        wh_id = self.warehouse_name_to_id.get(self.var_warehouse.get())
        if not wh_id:
            messagebox.showwarning("Atenção", "Escolhe um armazém.")
            return

        try:
            response = self.db.supabase.table('warehouse_stock').select('stock').eq('variant_id', self.loaded_variant_id).eq('warehouse_id', wh_id).execute()
            if response.data:
                current = response.data[0]['stock']
                messagebox.showinfo("Stock Atual", f"Armazém: {self.var_warehouse.get()}\nStock atual: {current}")
            else:
                messagebox.showinfo("Stock Atual", f"Armazém: {self.var_warehouse.get()}\nStock atual: 0 (sem registo)")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao consultar stock.\n\n{e}")

    def _refresh_stock_display(self):
        """Atualiza a visualização de stock após alterações"""
        if not self.loaded_variant_id:
            return
        
        try:
            full = self.db.get_full_view_by_variant_id(self.loaded_variant_id)
            if not full:
                return
            
            header, stocks = full
            
            self.txt_stock_info.config(state="normal")
            self.txt_stock_info.delete("1.0", tk.END)
            
            if stocks:
                self.txt_stock_info.insert(tk.END, "Stock disponível nos armazéns:\n\n")
                total_stock = 0
                for s in stocks:
                    self.txt_stock_info.insert(tk.END, f"  • {s['armazem']}: {s['stock']} unidades\n")
                    total_stock += s['stock']
                self.txt_stock_info.insert(tk.END, f"\n📦 Total geral: {total_stock} unidades")
            else:
                self.txt_stock_info.insert(tk.END, "⚠️ Nenhum stock registado em armazéns.\n\n")
                self.txt_stock_info.insert(tk.END, "Use os botões 'Adicionar' para criar stock.")
            
            self.txt_stock_info.config(state="disabled")
        except Exception as e:
            pass  # Silenciosamente ignora erros na atualização

    def bulk_update_stock(self):
        """Abre janela para alteração em massa de stock"""
        BulkUpdateWindow(self, self.app)


class BulkUpdateWindow(tk.Toplevel):
    """Janela para alteração em massa de stock"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self.db = app.db
        self.title("Alteração em Massa de Stock")
        self.geometry("950x750")
        self.resizable(True, True)
        
        # Frame principal
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="Alteração em Massa de Stock", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 12))
        
        # Opção de busca
        ttk.Label(main_frame, text="Adicionar Produtos:", font=("Segoe UI", 10, "bold")).grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(6, 4))
        
        ttk.Label(main_frame, text="Buscar por:").grid(row=2, column=0, sticky="w")
        self.var_search_type = tk.StringVar(value="ref_keyinvoice")
        
        search_frame = ttk.Frame(main_frame)
        search_frame.grid(row=2, column=1, columnspan=2, sticky="w")
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type, 
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type, 
                       value="ref_woocommerce").pack(side="left")
        
        ttk.Label(main_frame, text="Código:").grid(row=3, column=0, sticky="w", pady=(8, 0))
        self.entry_code = ttk.Entry(main_frame, width=30)
        self.entry_code.grid(row=3, column=1, sticky="w", pady=(8, 0))
        self.entry_code.bind("<Return>", lambda e: self.add_product_by_code())
        
        btn_frame1 = ttk.Frame(main_frame)
        btn_frame1.grid(row=3, column=2, columnspan=2, sticky="w", pady=(8, 0), padx=(8, 0))
        ttk.Button(btn_frame1, text="Adicionar", command=self.add_product_by_code).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame1, text="Carregar Excel", command=self.load_from_excel).pack(side="left")
        ttk.Button(btn_frame1, text="Limpar Lista", command=self.clear_products).pack(side="left", padx=(8, 0))
        
        # Tabela de produtos verificados
        ttk.Label(main_frame, text="Produtos a Alterar:", font=("Segoe UI", 10, "bold")).grid(
            row=4, column=0, columnspan=4, sticky="w", pady=(8, 4))
        
        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "stock")
        self.tree_products = ttk.Treeview(main_frame, columns=cols, show="headings", height=8)
        self.tree_products.grid(row=5, column=0, columnspan=4, sticky="nsew", pady=(0, 8))
        
        self.tree_products.heading("gtin", text="GTIN")
        self.tree_products.heading("modelo", text="Modelo")
        self.tree_products.heading("marca", text="Marca")
        self.tree_products.heading("cor", text="Cor")
        self.tree_products.heading("tamanho", text="Tamanho")
        self.tree_products.heading("stock", text="Stock Atual")
        
        self.tree_products.column("gtin", width=100, anchor="w")
        self.tree_products.column("modelo", width=120, anchor="w")
        self.tree_products.column("marca", width=80, anchor="w")
        self.tree_products.column("cor", width=70, anchor="w")
        self.tree_products.column("tamanho", width=60, anchor="w")
        self.tree_products.column("stock", width=150, anchor="w")
        
        tree_scroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree_products.yview)
        tree_scroll.grid(row=5, column=4, sticky="ns", pady=(0, 8))
        self.tree_products.config(yscrollcommand=tree_scroll.set)
        
        ttk.Separator(main_frame, orient="horizontal").grid(row=6, column=0, columnspan=5, sticky="ew", pady=8)
        
        # Configurações da operação
        ttk.Label(main_frame, text="Configuração da Operação:", font=("Segoe UI", 10, "bold")).grid(
            row=7, column=0, columnspan=3, sticky="w", pady=(0, 8))
        
        # Armazém
        ttk.Label(main_frame, text="Armazém:").grid(row=8, column=0, sticky="w", pady=4)
        self.var_warehouse = tk.StringVar()
        self.combo_warehouse = ttk.Combobox(main_frame, textvariable=self.var_warehouse, width=30, state="readonly")
        self.combo_warehouse.grid(row=8, column=1, columnspan=2, sticky="w", pady=4)
        
        # Tipo de operação
        ttk.Label(main_frame, text="Operação:").grid(row=9, column=0, sticky="w", pady=4)
        self.var_operation = tk.StringVar(value="add")
        
        op_frame = ttk.Frame(main_frame)
        op_frame.grid(row=9, column=1, columnspan=2, sticky="w", pady=4)
        ttk.Radiobutton(op_frame, text="Adicionar", variable=self.var_operation, value="add").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(op_frame, text="Remover", variable=self.var_operation, value="remove").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(op_frame, text="Definir", variable=self.var_operation, value="set").pack(side="left")
        
        # Quantidade
        ttk.Label(main_frame, text="Quantidade:").grid(row=10, column=0, sticky="w", pady=4)
        self.quantity = ttk.Entry(main_frame, width=15)
        self.quantity.grid(row=10, column=1, sticky="w", pady=4)
        
        ttk.Label(main_frame, text="(Para 'Adicionar' e 'Remover', usa a mesma quantidade para todos)", 
                 font=("Segoe UI", 8), foreground="gray").grid(row=11, column=0, columnspan=3, sticky="w", pady=(0, 8))
        
        ttk.Separator(main_frame, orient="horizontal").grid(row=12, column=0, columnspan=5, sticky="ew", pady=8)
        
        # Botões de ação
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=13, column=0, columnspan=3, pady=10)
        
        ttk.Button(btn_frame, text="Processar", command=self.process_bulk_update, width=15).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy, width=15).pack(side="left", padx=5)
        
        # Log de resultados
        ttk.Label(main_frame, text="Resultado:", font=("Segoe UI", 10, "bold")).grid(
            row=14, column=0, columnspan=3, sticky="w", pady=(8, 4))
        
        self.txt_log = tk.Text(main_frame, height=10, width=50, state="disabled", wrap="word")
        self.txt_log.grid(row=15, column=0, columnspan=3, sticky="nsew", pady=(0, 8))
        
        log_scroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.txt_log.yview)
        log_scroll.grid(row=15, column=3, sticky="ns", pady=(0, 8))
        self.txt_log.config(yscrollcommand=log_scroll.set)
        
        # Configurar grid
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)
        main_frame.rowconfigure(15, weight=1)
        
        # Carregar armazéns
        self._load_warehouses()
        
    def _load_warehouses(self):
        """Carrega lista de armazéns"""
        try:
            response = self.db.supabase.table('warehouses').select('id, name').execute()
            warehouses = {row['name']: row['id'] for row in response.data}
            self.warehouse_name_to_id = warehouses
            self.combo_warehouse['values'] = list(warehouses.keys())
            if self.combo_warehouse['values']:
                self.var_warehouse.set(self.combo_warehouse['values'][0])
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar armazéns.\n\n{e}")
    
    def load_from_excel(self):
        """Carrega lista de GTINs de um arquivo Excel"""
        file_path = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            
            gtins = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Pular header
                if row[0]:  # Assume que GTIN está na primeira coluna
                    gtin = str(row[0]).strip()
                    if gtin:
                        gtins.append(gtin)
            
            # Adicionar produtos do Excel à lista
            found_count = 0
            not_found = []
            
            for gtin in gtins:
                success, msg = self._add_product_to_list(gtin, "gtin")
                if success:
                    found_count += 1
                else:
                    not_found.append(f"{gtin}: {msg}")
            
            msg = f"{found_count} produtos adicionados do Excel."
            if not_found:
                msg += f"\n\nNão encontrados:\n" + "\n".join(not_found[:10])
                if len(not_found) > 10:
                    msg += f"\n... e mais {len(not_found) - 10}"
            
            messagebox.showinfo("Carregamento Concluído", msg)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar arquivo Excel.\n\n{e}")
    
    def add_product_by_code(self):
        """Adiciona produto por código Ref KeyInvoice ou Ref WooCommerce"""
        code = self.entry_code.get().strip()
        if not code:
            messagebox.showwarning("Atenção", "Informa o código.")
            return
        
        search_type = self.var_search_type.get()
        success, msg = self._add_product_to_list(code, search_type)
        
        if success:
            self.entry_code.delete(0, tk.END)
            self.entry_code.focus()
        else:
            messagebox.showwarning("Não encontrado", msg)
    
    def _add_product_to_list(self, code, search_type):
        """Busca produto e adiciona à lista. Retorna (success, message)"""
        try:
            # Buscar produtos pelo código
            results = self.db.search_variants(code, search_type)
            
            if not results:
                return False, f"Código '{code}' não encontrado"
            
            # Adicionar todos os resultados à tabela
            added = 0
            for result in results:
                variant_id = result.get("variant_id")
                gtin = result.get("gtin") or ""
                modelo = result.get("nome_modelo") or "N/A"
                marca = result.get("marca") or "N/A"
                cor = result.get("cor") or "N/A"
                tamanho = result.get("tamanho") or "N/A"
                
                # Verificar se já existe na tabela
                already_exists = False
                for item in self.tree_products.get_children():
                    if self.tree_products.item(item, "values")[0] == gtin:
                        already_exists = True
                        break
                
                if already_exists:
                    continue
                
                # Buscar stock de todos os armazéns
                stock_response = self.db.supabase.table('warehouse_stock').select(
                    'stock, warehouses(name)'
                ).eq('variant_id', variant_id).execute()
                
                if stock_response.data:
                    stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                    stock_text = " | ".join(stocks)
                else:
                    stock_text = "Sem stock"
                
                # Adicionar à tabela
                self.tree_products.insert(
                    "", "end",
                    values=(gtin, modelo, marca, cor, tamanho, stock_text),
                    tags=(variant_id,)
                )
                added += 1
            
            if added == 0:
                return False, "Produto já está na lista"
            
            return True, f"{added} produto(s) adicionado(s)"
            
        except Exception as e:
            return False, f"Erro: {str(e)}"
    
    def clear_products(self):
        """Limpa a lista de produtos"""
        for item in self.tree_products.get_children():
            self.tree_products.delete(item)
    
    def log(self, message):
        """Adiciona mensagem ao log"""
        self.txt_log.config(state="normal")
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state="disabled")
        self.update()
    
    def process_bulk_update(self):
        """Processa alteração em massa"""
        # Validar se há produtos na tabela
        if not self.tree_products.get_children():
            messagebox.showwarning("Atenção", "Adiciona pelo menos um produto à lista.")
            return
        
        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém.")
            return
        
        operation = self.var_operation.get()
        quantity_str = self.quantity.get().strip()
        
        if not quantity_str:
            messagebox.showwarning("Atenção", "Informa a quantidade.")
            return
        
        try:
            quantity = int(quantity_str)
            if quantity <= 0:
                messagebox.showwarning("Atenção", "Quantidade deve ser maior que 0.")
                return
        except ValueError:
            messagebox.showwarning("Atenção", "Quantidade inválida.")
            return
        
        wh_id = self.warehouse_name_to_id.get(wh_name)
        
        # Contar produtos na tabela
        products_count = len(self.tree_products.get_children())
        
        if not messagebox.askyesno("Confirmar", 
                                   f"Confirma alteração em massa?\n\n"
                                   f"Produtos: {products_count}\n"
                                   f"Armazém: {wh_name}\n"
                                   f"Operação: {operation}\n"
                                   f"Quantidade: {quantity}"):
            return
        
        # Limpar log
        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", tk.END)
        self.txt_log.config(state="disabled")
        
        self.log(f"Iniciando alteração em massa...")
        self.log(f"Total de produtos: {products_count}\n")
        
        success_count = 0
        error_count = 0
        
        # Processar cada produto da tabela
        for item in self.tree_products.get_children():
            values = self.tree_products.item(item, "values")
            gtin = values[0]
            modelo = values[1]
            marca = values[2]
            cor = values[3]
            tamanho = values[4]
            variant_id = self.tree_products.item(item, "tags")[0]
            
            produto_info = f"{gtin} - {modelo} | {marca} | {cor} | Tam: {tamanho}"
            
            try:
                # Executar operação
                if operation == "add":
                    success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, quantity)
                elif operation == "remove":
                    success, msg = self.app.product_service.remove_from_stock(variant_id, wh_id, quantity)
                elif operation == "set":
                    # Para "definir", precisamos verificar o stock atual e ajustar
                    current_response = self.db.supabase.table('warehouse_stock').select('stock').eq(
                        'variant_id', variant_id).eq('warehouse_id', wh_id).execute()
                    
                    if current_response.data:
                        current_stock = current_response.data[0]['stock']
                        diff = quantity - current_stock
                        
                        if diff > 0:
                            success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, diff)
                        elif diff < 0:
                            success, msg = self.app.product_service.remove_from_stock(variant_id, wh_id, abs(diff))
                        else:
                            success, msg = True, f"Stock já está em {quantity}"
                    else:
                        # Criar novo registro
                        success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, quantity)
                
                if success:
                    self.log(f"✓ {produto_info}")
                    self.log(f"  → {msg}\n")
                    success_count += 1
                    
                    # Auditoria
                    self.app.db.audit(self.app.user, f"BULK_{operation.upper()}_STOCK", "warehouse_stock",
                                    entity_pk=f"variant_id={variant_id},warehouse_id={wh_id}",
                                    details={"gtin": gtin, "quantity": quantity, "operation": operation})
                else:
                    self.log(f"❌ {produto_info}")
                    self.log(f"  → {msg}\n")
                    error_count += 1
                    
            except Exception as e:
                self.log(f"❌ GTIN {gtin}: Erro - {str(e)}\n")
                error_count += 1
        
        self.log(f"\n{'='*50}")
        self.log(f"Processamento concluído!")
        self.log(f"Sucesso: {success_count}")
        self.log(f"Erros: {error_count}")
        self.log(f"Total: {products_count}")
        
        messagebox.showinfo("Concluído", 
                          f"Alteração em massa concluída!\n\n"
                          f"Sucesso: {success_count}\n"
                          f"Erros: {error_count}\n"
                          f"Total: {products_count}")


class DeleteTab(BaseTab):
    """Tab para excluir produtos"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, app)
        self.load_domains()

        ttk.Label(self, text="Excluir Produto (por Código + Armazém)", 
                 font=("Segoe UI", 14, "bold")).grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 12))

        ttk.Label(self, text="Buscar por:").grid(row=1, column=0, sticky="w")
        self.var_search_type = tk.StringVar(value="gtin")
        
        search_frame = ttk.Frame(self)
        search_frame.grid(row=1, column=1, columnspan=3, sticky="w")
        
        ttk.Radiobutton(search_frame, text="GTIN", variable=self.var_search_type, 
                       value="gtin").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type, 
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type, 
                       value="ref_woocommerce").pack(side="left")

        ttk.Label(self, text="Código:").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.search_value = ttk.Entry(self, width=28)
        self.search_value.grid(row=2, column=1, sticky="w", padx=(0, 12), pady=(8, 0))

        ttk.Label(self, text="Armazém:").grid(row=2, column=2, sticky="w", pady=(8, 0))
        self.var_warehouse = tk.StringVar()
        self.combo_warehouse = ttk.Combobox(self, textvariable=self.var_warehouse, width=25, state="readonly")
        self.combo_warehouse.grid(row=2, column=3, sticky="w", pady=(8, 0))

        self._refresh_warehouses()

        ttk.Button(self, text="Carregar", command=self.preview).grid(row=2, column=4, padx=(8, 0), pady=(8, 0))
        ttk.Separator(self).grid(row=3, column=0, columnspan=6, sticky="ew", pady=12)

        self.txt = tk.Text(self, height=18, width=110)
        self.txt.grid(row=4, column=0, columnspan=6, sticky="nsew")
        
        # Menu contexto para copy/paste
        self.txt_menu = tk.Menu(self.txt, tearoff=0)
        self.txt_menu.add_command(label="Copiar", command=lambda: copy_to_clipboard(self.txt))
        self.txt_menu.add_command(label="Colar", command=lambda: paste_from_clipboard(self.txt))
        self.txt.bind("<Button-3>", lambda e: self.txt_menu.post(e.x_root, e.y_root))
        self.txt.bind("<Control-c>", lambda e: copy_to_clipboard(self.txt))

        btns = ttk.Frame(self)
        btns.grid(row=5, column=0, columnspan=6, sticky="w", pady=(10, 0))
        ttk.Button(btns, text="Excluir stock deste armazém", command=self.delete_stock).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="Cancelar / Limpar", command=self.clear).grid(row=0, column=1)

        self.loaded = None
        self.grid_rowconfigure(4, weight=1)
        self.grid_columnconfigure(5, weight=1)

    def _refresh_warehouses(self):
        """Atualiza lista de armazéns"""
        whs = self.domain_service.get_domain_list("warehouses")
        self.wh_id_to_name = {r[0]: r[1] for r in whs}
        self.wh_name_to_id = {r[1]: r[0] for r in whs}
        self.combo_warehouse["values"] = list(self.wh_name_to_id.keys())
        if self.combo_warehouse["values"]:
            self.var_warehouse.set(self.combo_warehouse["values"][0])

    def clear(self):
        """Limpa formulário"""
        self.loaded = None
        self.txt.delete("1.0", tk.END)
        self.search_value.delete(0, tk.END)

    def preview(self):
        """Carrega preview do produto"""
        value = self.search_value.get().strip()
        search_type = self.var_search_type.get()
        
        if not value:
            messagebox.showwarning("Atenção", "Informa o código.")
            return
        
        try:
            full = self.db.get_full_view_by_gtin(value, search_type)
            self.txt.delete("1.0", tk.END)
            
            if not full:
                self.txt.insert(tk.END, "Código não encontrado.\n")
                self.loaded = None
                return

            header, stocks = full
            self.loaded = header

            self.txt.insert(tk.END, f"GTIN: {header['gtin']}\n")
            self.txt.insert(tk.END, f"Ref KeyInvoice: {header['ref_keyinvoice']}\n")
            self.txt.insert(tk.END, f"Ref WooCommerce: {header['ref_woocomerce']}\n")
            self.txt.insert(tk.END, f"Modelo: {header['nome_modelo']}\n")
            self.txt.insert(tk.END, f"Marca: {header['marca']}\n")
            self.txt.insert(tk.END, f"Categoria/Subcategoria: {header['categoria']} / {header['subcategoria']}\n")
            self.txt.insert(tk.END, f"Fornecedor: {header['fornecedor']}\n")
            self.txt.insert(tk.END, f"Cor/Tamanho: {header['cor']} / {header['tamanho']}\n\n")

            self.txt.insert(tk.END, "Stock por armazém:\n")
            for s in stocks:
                self.txt.insert(tk.END, f" - {s['armazem']}: {s['stock']}\n")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha.\n\n{e}")

    def delete_stock(self):
        """Apaga stock"""
        if not self.loaded:
            messagebox.showwarning("Atenção", "Carrega um produto primeiro.")
            return
            
        wh_name = self.var_warehouse.get()
        wh_id = self.wh_name_to_id.get(wh_name)
        if not wh_id:
            messagebox.showwarning("Atenção", "Escolhe armazém.")
            return

        variant_id = self.loaded["variant_id"]
        gtin = self.loaded["gtin"]

        if not messagebox.askyesno("Confirmar", 
                                   f"Excluir o stock do GTIN {gtin} no armazém '{wh_name}'?"):
            return

        try:
            success, msg, deleted_variant = self.app.product_service.delete_stock(variant_id, wh_id)
            
            self.app.db.audit(self.app.user, "DELETE_STOCK", "warehouse_stock",
                            entity_pk=f"variant_id={variant_id},warehouse_id={wh_id}",
                            details={"gtin": gtin, "warehouse": wh_name})

            if deleted_variant:
                self.app.db.audit(self.app.user, "DELETE_VARIANT", "product_variant",
                                entity_pk=f"id={variant_id}",
                                details={"gtin": gtin, "reason": "no_stock_rows"})

            messagebox.showinfo("OK", msg)
            self.preview()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao excluir.\n\n{e}")


class ViewTab(BaseTab):
    """Tab para visualizar produtos"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, app)

        ttk.Label(self, text="Visualizar Produto", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12))

        ttk.Label(self, text="Buscar por:").grid(row=1, column=0, sticky="w")
        self.var_search_type = tk.StringVar(value="gtin")
        
        search_frame = ttk.Frame(self)
        search_frame.grid(row=1, column=1, columnspan=3, sticky="w")
        
        ttk.Radiobutton(search_frame, text="GTIN", variable=self.var_search_type, 
                       value="gtin").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type, 
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type, 
                       value="ref_woocommerce").pack(side="left")

        ttk.Label(self, text="Código:").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.search_value = ttk.Entry(self, width=40)
        self.search_value.grid(row=2, column=1, sticky="w", padx=(0, 8), pady=(8, 0))

        ttk.Button(self, text="Buscar", command=self.search).grid(row=2, column=2, sticky="w", pady=(8, 0))
        ttk.Button(self, text="Limpar", command=self.clear).grid(row=2, column=3, sticky="w", pady=(8, 0))
        ttk.Button(self, text="Descarregar Excel", command=self.export_to_excel).grid(row=2, column=4, sticky="w", padx=(8, 0), pady=(8, 0))

        ttk.Separator(self).grid(row=3, column=0, columnspan=6, sticky="ew", pady=12)

        ttk.Label(self, text="Resultados", font=("Segoe UI", 10, "bold")).grid(row=4, column=0, columnspan=6, sticky="w")

        # Treeview para resultados
        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "ref_keyinvoice", "ref_woocommerce", "stock")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=8)
        self.tree.grid(row=5, column=0, columnspan=6, sticky="nsew", pady=(6, 12))

        self.tree.heading("gtin", text="GTIN")
        self.tree.heading("modelo", text="Modelo")
        self.tree.heading("marca", text="Marca")
        self.tree.heading("cor", text="Cor")
        self.tree.heading("tamanho", text="Tamanho")
        self.tree.heading("ref_keyinvoice", text="Ref KeyInvoice")
        self.tree.heading("ref_woocommerce", text="Ref WooCommerce")
        self.tree.heading("stock", text="Stock (Armazéns)")

        self.tree.column("gtin", width=85, anchor="w")
        self.tree.column("modelo", width=100, anchor="w")
        self.tree.column("marca", width=70, anchor="w")
        self.tree.column("cor", width=60, anchor="w")
        self.tree.column("tamanho", width=60, anchor="w")
        self.tree.column("ref_keyinvoice", width=90, anchor="w")
        self.tree.column("ref_woocommerce", width=90, anchor="w")
        self.tree.column("stock", width=180, anchor="w")
        
        # Double click para ver detalhes
        self.tree.bind("<Double-1>", lambda e: self.show_details())
        
        # Menu contexto
        self.tree_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_menu.add_command(label="Ver Detalhes", command=self.show_details)
        self.tree.bind("<Button-3>", lambda e: self.tree_menu.post(e.x_root, e.y_root))

        ttk.Label(self, text="Detalhes", font=("Segoe UI", 10, "bold")).grid(row=6, column=0, columnspan=6, sticky="w", pady=(12, 6))

        self.txt = tk.Text(self, height=8, width=110)
        self.txt.grid(row=7, column=0, columnspan=6, sticky="nsew")
        
        # Menu contexto para copy/paste
        self.txt_menu = tk.Menu(self.txt, tearoff=0)
        self.txt_menu.add_command(label="Copiar", command=lambda: copy_to_clipboard(self.txt))
        self.txt_menu.add_command(label="Colar", command=lambda: paste_from_clipboard(self.txt))
        self.txt.bind("<Button-3>", lambda e: self.txt_menu.post(e.x_root, e.y_root))
        self.txt.bind("<Control-c>", lambda e: copy_to_clipboard(self.txt))

        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(7, weight=1)
        
        self.selected_variant_id = None

    def export_to_excel(self):
        """Exporta resultados da busca para Excel"""
        if not self.tree.get_children():
            messagebox.showwarning("Atenção", "Não há resultados para exportar.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
            
            # Header
            headers = ["GTIN", "Modelo", "Marca", "Cor", "Tamanho", "Ref KeyInvoice", "Ref WooCommerce", "Stock (Armazéns)"]
            ws.append(headers)
            
            # Estilo header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dados
            for item in self.tree.get_children():
                values = self.tree.item(item, "values")
                ws.append(values)
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 25
            
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Relatório exportado para:\n{file_path}")
            
            self.app.db.audit(self.app.user, "EXPORT_EXCEL", "products", 
                            details={"file": file_path, "rows": len(self.tree.get_children())})
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar.\n\n{e}")

    def clear(self):
        """Limpa formulário"""
        self.search_value.delete(0, tk.END)
        self.txt.delete("1.0", tk.END)
        for item in self.tree.get_children():
            self.tree.delete(item)

    def search(self):
        """Pesquisa produtos"""
        value = self.search_value.get().strip()
        search_type = self.var_search_type.get()
        
        if not value:
            messagebox.showwarning("Atenção", "Informa o código de busca.")
            return
        
        try:
            # Limpar tabela e detalhes
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.txt.delete("1.0", tk.END)
            
            # Buscar variantes
            results = self.db.search_variants(value, search_type)
            
            if not results:
                tipo_busca = {
                    'gtin': 'GTIN',
                    'ref_keyinvoice': 'Ref KeyInvoice',
                    'ref_woocommerce': 'Ref WooCommerce'
                }
                self.txt.insert(tk.END, f"{tipo_busca[search_type]} não encontrado.\n")
                return

            # Preencher tabela com stock
            for r in results:
                variant_id = r["variant_id"]
                
                # Buscar stock dos armazéns
                stock_response = self.db.supabase.table('warehouse_stock').select(
                    'stock, warehouses(name)'
                ).eq('variant_id', variant_id).execute()
                
                stock_text = ""
                if stock_response.data:
                    stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                    stock_text = " | ".join(stocks)
                else:
                    stock_text = "Sem stock"
                
                self.tree.insert(
                    "", "end",
                    values=(
                        r.get("gtin") or "",
                        r.get("nome_modelo") or "",
                        r.get("marca") or "",
                        r.get("cor") or "",
                        r.get("tamanho") or "",
                        r.get("ref_keyinvoice") or "",
                        r.get("ref_woocomerce") or "",
                        stock_text
                    ),
                    tags=(r["variant_id"],)
                )
            
            if len(results) == 1:
                # Auto-selecionar se houver apenas um resultado
                self.tree.selection_set(self.tree.get_children()[0])
                self.show_details()
            else:
                messagebox.showinfo("Encontrado", f"Encontradas {len(results)} variação(ões).\n\nDouble-click para ver detalhes.")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Falha na busca.\n\n{e}")

    def show_details(self):
        """Mostra detalhes do produto selecionado"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Seleciona uma variação na tabela.")
            return
        
        # Obter variant_id da tag
        tags = self.tree.item(sel[0], "tags")
        if not tags:
            messagebox.showwarning("Atenção", "Erro ao obter dados da variação.")
            return
        
        variant_id = tags[0]
        
        try:
            full = self.db.get_full_view_by_variant_id(variant_id)
            self.txt.delete("1.0", tk.END)
            
            if not full:
                self.txt.insert(tk.END, "Variação não encontrada.\n")
                return

            header, stocks = full
            self.selected_variant_id = variant_id

            self.txt.insert(tk.END, f"GTIN: {header['gtin']}\n")
            self.txt.insert(tk.END, f"Ref KeyInvoice: {header['ref_keyinvoice']}\n")
            self.txt.insert(tk.END, f"Ref WooCommerce: {header['ref_woocomerce']}\n")
            self.txt.insert(tk.END, f"Modelo: {header['nome_modelo']}\n")
            self.txt.insert(tk.END, f"Marca: {header['marca']}\n")
            self.txt.insert(tk.END, f"Categoria/Subcategoria: {header['categoria']} / {header['subcategoria']}\n")
            self.txt.insert(tk.END, f"Fornecedor: {header['fornecedor']}\n")
            self.txt.insert(tk.END, f"Cor/Tamanho: {header['cor']} / {header['tamanho']}\n\n")

            self.txt.insert(tk.END, "Stock por armazém:\n")
            for s in stocks:
                self.txt.insert(tk.END, f" - {s['armazem']}: {s['stock']}\n")

            self.app.db.audit(
                self.app.user, 
                "VIEW_DETAILS", 
                "product_variant", 
                entity_pk=f"id={variant_id}", 
                details={}
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar detalhes.\n\n{e}")


def main():
    """Função principal"""
    db = DB()
    app = App(db)
    app.mainloop()


class WarehouseTab(BaseTab):
    """Tab para visualizar produtos por armazém"""
    
    def __init__(self, parent, app: App):
        super().__init__(parent, app)
        self.load_domains()

        ttk.Label(self, text="Consultar Armazém", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12)
        )

        ttk.Label(self, text="Escolhe o Armazém:").grid(row=1, column=0, sticky="w")
        self.var_warehouse = tk.StringVar()
        self.combo_warehouse = ttk.Combobox(self, textvariable=self.var_warehouse, width=40, state="readonly")
        self.combo_warehouse.grid(row=1, column=1, columnspan=2, sticky="w", padx=(0, 8))
        
        ttk.Button(self, text="Carregar", command=self.load_warehouse).grid(row=1, column=3, sticky="w")
        ttk.Button(self, text="Atualizar Lista", command=self.refresh_warehouses).grid(row=1, column=4, sticky="w", padx=(8, 0))
        ttk.Button(self, text="Descarregar Excel", command=self.export_to_excel).grid(row=1, column=5, sticky="w", padx=(8, 0))

        self._refresh_warehouses()

        ttk.Separator(self).grid(row=2, column=0, columnspan=6, sticky="ew", pady=12)

        # Informação do armazém selecionado
        self.lbl_warehouse_info = ttk.Label(self, text="", font=("Segoe UI", 10, "bold"))
        self.lbl_warehouse_info.grid(row=3, column=0, columnspan=6, sticky="w")

        # Treeview para produtos do armazém
        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "quantidade")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=15)
        self.tree.grid(row=4, column=0, columnspan=6, sticky="nsew", pady=(12, 12))

        self.tree.heading("gtin", text="GTIN")
        self.tree.heading("modelo", text="Modelo")
        self.tree.heading("marca", text="Marca")
        self.tree.heading("cor", text="Cor")
        self.tree.heading("tamanho", text="Tamanho")
        self.tree.heading("quantidade", text="Quantidade")

        self.tree.column("gtin", width=100, anchor="w")
        self.tree.column("modelo", width=140, anchor="w")
        self.tree.column("marca", width=100, anchor="w")
        self.tree.column("cor", width=80, anchor="w")
        self.tree.column("tamanho", width=80, anchor="w")
        self.tree.column("quantidade", width=100, anchor="center")

        # Menu contexto
        self.tree_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_menu.add_command(label="Copiar GTIN", command=self.copy_gtin)
        self.tree.bind("<Button-3>", lambda e: self.tree_menu.post(e.x_root, e.y_root))

        # Resumo
        self.lbl_summary = ttk.Label(self, text="", font=("Segoe UI", 9))
        self.lbl_summary.grid(row=5, column=0, columnspan=6, sticky="w")

        self.grid_rowconfigure(4, weight=1)

    def _refresh_warehouses(self):
        """Atualiza lista de armazéns"""
        whs = self.domain_service.get_domain_list("warehouses")
        self.wh_id_to_name = {r[0]: r[1] for r in whs}
        self.wh_name_to_id = {r[1]: r[0] for r in whs}
        self.combo_warehouse["values"] = list(self.wh_name_to_id.keys())
        if self.combo_warehouse["values"]:
            self.var_warehouse.set(self.combo_warehouse["values"][0])

    def refresh_warehouses(self):
        """Recarrega lista de armazéns"""
        self._refresh_warehouses()
        messagebox.showinfo("OK", "Lista de armazéns atualizada.")

    def load_warehouse(self):
        """Carrega produtos do armazém selecionado"""
        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém.")
            return
        
        wh_id = self.wh_name_to_id.get(wh_name)
        
        try:
            # Limpar tabela
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Buscar todos os produtos deste armazém
            response = self.db.supabase.table('warehouse_stock').select(
                'stock, product_variant(id, gtin, product_model(nome_modelo, brands(name), categories(name), subcategories(name)), colors(name), sizes(value))'
            ).eq('warehouse_id', wh_id).order('product_variant(gtin)').execute()

            if not response.data:
                self.lbl_warehouse_info.config(text=f"Armazém: {wh_name} - Sem produtos")
                self.lbl_summary.config(text="")
                return

            # Preencher tabela
            total_items = 0
            for row in response.data:
                variant = row['product_variant']
                stock = row['stock']
                total_items += stock

                self.tree.insert(
                    "", "end",
                    values=(
                        variant.get('gtin') or "",
                        variant['product_model']['nome_modelo'] if variant.get('product_model') else "",
                        variant['product_model']['brands']['name'] if variant.get('product_model') and variant['product_model'].get('brands') else "",
                        variant['colors']['name'] if variant.get('colors') else "",
                        variant['sizes']['value'] if variant.get('sizes') else "",
                        stock
                    ),
                    tags=(variant['id'],)
                )

            self.lbl_warehouse_info.config(text=f"Armazém: {wh_name} - {len(response.data)} variação(ões)")
            self.lbl_summary.config(text=f"Total de itens em stock: {total_items}")

            self.app.db.audit(
                self.app.user,
                "VIEW_WAREHOUSE",
                "warehouse_stock",
                entity_pk=f"warehouse_id={wh_id}",
                details={"warehouse_name": wh_name}
            )

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar armazém.\n\n{e}")

    def copy_gtin(self):
        """Copia GTIN do item selecionado"""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Seleciona um produto.")
            return
        
        gtin = self.tree.item(sel[0], "values")[0]
        self.tree.clipboard_clear()
        self.tree.clipboard_append(gtin)
        self.tree.update()
        messagebox.showinfo("OK", f"GTIN {gtin} copiado para clipboard!")

    def export_to_excel(self):
        """Exporta lista do armazém para Excel"""
        if not self.tree.get_children():
            messagebox.showwarning("Atenção", "Não há resultados para exportar.")
            return
        
        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém primeiro.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"armazem_{wh_name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = wh_name[:31]  # Limitar nome da sheet a 31 caracteres
            
            # Header
            headers = ["GTIN", "Modelo", "Marca", "Cor", "Tamanho", "Quantidade"]
            ws.append(headers)
            
            # Estilo header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Dados
            total_items = 0
            for item in self.tree.get_children():
                values = self.tree.item(item, "values")
                ws.append(values)
                total_items += int(values[5]) if values[5] else 0
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            
            # Adicionar resumo
            ws.append([])
            ws.append(["Total de itens", total_items])
            resumo_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            resumo_font = Font(bold=True)
            for cell in ws[len(ws)]:
                cell.fill = resumo_fill
                cell.font = resumo_font
            
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Relatório exportado para:\n{file_path}")
            
            self.app.db.audit(self.app.user, "EXPORT_WAREHOUSE_EXCEL", "warehouse_stock",
                            details={"warehouse": wh_name, "file": file_path, "rows": len(self.tree.get_children())})
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar.\n\n{e}")





if __name__ == "__main__":
    from ui.main import main as ui_main
    ui_main()
