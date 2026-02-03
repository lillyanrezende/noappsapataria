from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ui.components.helpers import BaseTab, copy_to_clipboard, paste_from_clipboard


class ViewTab(BaseTab):
    """Tab para visualizar produtos"""

    def __init__(self, parent, app):
        super().__init__(parent, app)

        ttk.Label(self, text="Visualizar Produto", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12))

        ttk.Label(self, text="Buscar por:").grid(row=1, column=0, sticky="w")
        self.var_search_type = tk.StringVar(value="gtin")

        search_frame = ttk.Frame(self)
        search_frame.grid(row=1, column=1, columnspan=3, sticky="w")

        ttk.Radiobutton(search_frame, text="GTIN", variable=self.var_search_type,
                       value="gtin").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type,
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type,
                       value="ref_woocommerce").pack(side="left")

        ttk.Label(self, text="Código:").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.search_value = ttk.Entry(self, width=40)
        self.search_value.grid(row=2, column=1, sticky="w", padx=(0, 8), pady=(8, 0))

        ttk.Button(self, text="Buscar", command=self.search).grid(row=2, column=2, sticky="w", pady=(8, 0))
        ttk.Button(self, text="Limpar", command=self.clear).grid(row=2, column=3, sticky="w", pady=(8, 0))
        ttk.Button(self, text="Descarregar Excel", command=self.export_to_excel).grid(row=2, column=4, sticky="w", padx=(8, 0), pady=(8, 0))

        ttk.Separator(self).grid(row=3, column=0, columnspan=6, sticky="ew", pady=12)

        ttk.Label(self, text="Resultados", font=("Segoe UI", 10, "bold")).grid(row=4, column=0, columnspan=6, sticky="w")

        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "ref_keyinvoice", "ref_woocommerce", "stock")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=8)
        self.tree.grid(row=5, column=0, columnspan=6, sticky="nsew", pady=(6, 12))

        self.tree.heading("gtin", text="GTIN")
        self.tree.heading("modelo", text="Modelo")
        self.tree.heading("marca", text="Marca")
        self.tree.heading("cor", text="Cor")
        self.tree.heading("tamanho", text="Tamanho")
        self.tree.heading("ref_keyinvoice", text="Ref KeyInvoice")
        self.tree.heading("ref_woocommerce", text="Ref WooCommerce")
        self.tree.heading("stock", text="Stock (Armazéns)")

        self.tree.column("gtin", width=85, anchor="w")
        self.tree.column("modelo", width=100, anchor="w")
        self.tree.column("marca", width=70, anchor="w")
        self.tree.column("cor", width=60, anchor="w")
        self.tree.column("tamanho", width=60, anchor="w")
        self.tree.column("ref_keyinvoice", width=90, anchor="w")
        self.tree.column("ref_woocommerce", width=90, anchor="w")
        self.tree.column("stock", width=180, anchor="w")

        self.tree.bind("<Double-1>", lambda e: self.show_details())

        self.tree_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_menu.add_command(label="Ver Detalhes", command=self.show_details)
        self.tree.bind("<Button-3>", lambda e: self.tree_menu.post(e.x_root, e.y_root))

        ttk.Label(self, text="Detalhes", font=("Segoe UI", 10, "bold")).grid(row=6, column=0, columnspan=6, sticky="w", pady=(12, 6))

        self.txt = tk.Text(self, height=8, width=110)
        self.txt.grid(row=7, column=0, columnspan=6, sticky="nsew")

        self.txt_menu = tk.Menu(self.txt, tearoff=0)
        self.txt_menu.add_command(label="Copiar", command=lambda: copy_to_clipboard(self.txt))
        self.txt_menu.add_command(label="Colar", command=lambda: paste_from_clipboard(self.txt))
        self.txt.bind("<Button-3>", lambda e: self.txt_menu.post(e.x_root, e.y_root))
        self.txt.bind("<Control-c>", lambda e: copy_to_clipboard(self.txt))

        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(7, weight=1)

        self.selected_variant_id = None

    def export_to_excel(self):
        if not self.tree.get_children():
            messagebox.showwarning("Atenção", "Não há resultados para exportar.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"relatorio_produtos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"

            headers = ["GTIN", "Modelo", "Marca", "Cor", "Tamanho", "Ref KeyInvoice", "Ref WooCommerce", "Stock (Armazéns)"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for item in self.tree.get_children():
                values = self.tree.item(item, "values")
                ws.append(values)

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 25

            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Relatório exportado para:\n{file_path}")

            self.app.db.audit(self.app.user, "EXPORT_EXCEL", "products",
                            details={"file": file_path, "rows": len(self.tree.get_children())})
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar.\n\n{e}")

    def clear(self):
        self.search_value.delete(0, tk.END)
        self.txt.delete("1.0", tk.END)
        for item in self.tree.get_children():
            self.tree.delete(item)

    def search(self):
        value = self.search_value.get().strip()
        search_type = self.var_search_type.get()

        if not value:
            messagebox.showwarning("Atenção", "Informa o código de busca.")
            return

        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.txt.delete("1.0", tk.END)

            results = self.db.search_variants(value, search_type)

            if not results:
                tipo_busca = {
                    'gtin': 'GTIN',
                    'ref_keyinvoice': 'Ref KeyInvoice',
                    'ref_woocommerce': 'Ref WooCommerce'
                }
                self.txt.insert(tk.END, f"{tipo_busca[search_type]} não encontrado.\n")
                return

            for r in results:
                variant_id = r["variant_id"]

                stock_response = self.db.supabase.table('warehouse_stock').select(
                    'stock, warehouses(name)'
                ).eq('variant_id', variant_id).execute()

                if stock_response.data:
                    stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                    stock_text = " | ".join(stocks)
                else:
                    stock_text = "Sem stock"

                self.tree.insert(
                    "", "end",
                    values=(
                        r.get("gtin") or "",
                        r.get("nome_modelo") or "",
                        r.get("marca") or "",
                        r.get("cor") or "",
                        r.get("tamanho") or "",
                        r.get("ref_keyinvoice") or "",
                        r.get("ref_woocomerce") or "",
                        stock_text
                    ),
                    tags=(r["variant_id"],)
                )

            if len(results) == 1:
                self.tree.selection_set(self.tree.get_children()[0])
                self.show_details()
            else:
                messagebox.showinfo("Encontrado", f"Encontradas {len(results)} variação(ões).\n\nDouble-click para ver detalhes.")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha na busca.\n\n{e}")

    def show_details(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Seleciona uma variação na tabela.")
            return

        tags = self.tree.item(sel[0], "tags")
        if not tags:
            messagebox.showwarning("Atenção", "Erro ao obter dados da variação.")
            return

        variant_id = tags[0]

        try:
            full = self.db.get_full_view_by_variant_id(variant_id)
            self.txt.delete("1.0", tk.END)

            if not full:
                self.txt.insert(tk.END, "Variação não encontrada.\n")
                return

            header, stocks = full
            self.selected_variant_id = variant_id

            self.txt.insert(tk.END, f"GTIN: {header['gtin']}\n")
            self.txt.insert(tk.END, f"Ref KeyInvoice: {header['ref_keyinvoice']}\n")
            self.txt.insert(tk.END, f"Ref WooCommerce: {header['ref_woocomerce']}\n")
            self.txt.insert(tk.END, f"Modelo: {header['nome_modelo']}\n")
            self.txt.insert(tk.END, f"Marca: {header['marca']}\n")
            self.txt.insert(tk.END, f"Categoria/Subcategoria: {header['categoria']} / {header['subcategoria']}\n")
            self.txt.insert(tk.END, f"Fornecedor: {header['fornecedor']}\n")
            self.txt.insert(tk.END, f"Cor/Tamanho: {header['cor']} / {header['tamanho']}\n\n")

            self.txt.insert(tk.END, "Stock por armazém:\n")
            for s in stocks:
                self.txt.insert(tk.END, f" - {s['armazem']}: {s['stock']}\n")

            self.app.db.audit(
                self.app.user,
                "VIEW_DETAILS",
                "product_variant",
                entity_pk=f"id={variant_id}",
                details={}
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar detalhes.\n\n{e}")
