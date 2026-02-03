from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ui.components.helpers import BaseTab


class WarehouseTab(BaseTab):
    """Tab para visualizar produtos por armazém"""

    def __init__(self, parent, app):
        super().__init__(parent, app)
        self.load_domains()

        ttk.Label(self, text="Consultar Armazém", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=6, sticky="w", pady=(0, 12)
        )

        ttk.Label(self, text="Escolhe o Armazém:").grid(row=1, column=0, sticky="w")
        self.var_warehouse = tk.StringVar()
        self.combo_warehouse = ttk.Combobox(self, textvariable=self.var_warehouse, width=40, state="readonly")
        self.combo_warehouse.grid(row=1, column=1, columnspan=2, sticky="w", padx=(0, 8))

        ttk.Button(self, text="Carregar", command=self.load_warehouse).grid(row=1, column=3, sticky="w")
        ttk.Button(self, text="Atualizar Lista", command=self.refresh_warehouses).grid(row=1, column=4, sticky="w", padx=(8, 0))
        ttk.Button(self, text="Descarregar Excel", command=self.export_to_excel).grid(row=1, column=5, sticky="w", padx=(8, 0))

        self._refresh_warehouses()

        ttk.Separator(self).grid(row=2, column=0, columnspan=6, sticky="ew", pady=12)

        self.lbl_warehouse_info = ttk.Label(self, text="", font=("Segoe UI", 10, "bold"))
        self.lbl_warehouse_info.grid(row=3, column=0, columnspan=6, sticky="w")

        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "quantidade")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=15)
        self.tree.grid(row=4, column=0, columnspan=6, sticky="nsew", pady=(12, 12))

        self.tree.heading("gtin", text="GTIN")
        self.tree.heading("modelo", text="Modelo")
        self.tree.heading("marca", text="Marca")
        self.tree.heading("cor", text="Cor")
        self.tree.heading("tamanho", text="Tamanho")
        self.tree.heading("quantidade", text="Quantidade")

        self.tree.column("gtin", width=100, anchor="w")
        self.tree.column("modelo", width=140, anchor="w")
        self.tree.column("marca", width=100, anchor="w")
        self.tree.column("cor", width=80, anchor="w")
        self.tree.column("tamanho", width=80, anchor="w")
        self.tree.column("quantidade", width=100, anchor="center")

        self.tree_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_menu.add_command(label="Copiar GTIN", command=self.copy_gtin)
        self.tree.bind("<Button-3>", lambda e: self.tree_menu.post(e.x_root, e.y_root))

        self.lbl_summary = ttk.Label(self, text="", font=("Segoe UI", 9))
        self.lbl_summary.grid(row=5, column=0, columnspan=6, sticky="w")

        self.grid_rowconfigure(4, weight=1)

    def _refresh_warehouses(self):
        whs = self.domain_service.get_domain_list("warehouses")
        self.wh_id_to_name = {r[0]: r[1] for r in whs}
        self.wh_name_to_id = {r[1]: r[0] for r in whs}
        self.combo_warehouse["values"] = list(self.wh_name_to_id.keys())
        if self.combo_warehouse["values"]:
            self.var_warehouse.set(self.combo_warehouse["values"][0])

    def refresh_warehouses(self):
        self._refresh_warehouses()
        messagebox.showinfo("OK", "Lista de armazéns atualizada.")

    def load_warehouse(self):
        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém.")
            return

        wh_id = self.wh_name_to_id.get(wh_name)

        try:
            for item in self.tree.get_children():
                self.tree.delete(item)

            response = self.db.supabase.table('warehouse_stock').select(
                'stock, product_variant(id, gtin, product_model(nome_modelo, brands(name), categories(name), subcategories(name)), colors(name), sizes(value))'
            ).eq('warehouse_id', wh_id).order('product_variant(gtin)').execute()

            if not response.data:
                self.lbl_warehouse_info.config(text=f"Armazém: {wh_name} - Sem produtos")
                self.lbl_summary.config(text="")
                return

            total_items = 0
            for row in response.data:
                variant = row['product_variant']
                stock = row['stock']
                total_items += stock

                self.tree.insert(
                    "", "end",
                    values=(
                        variant.get('gtin') or "",
                        variant['product_model']['nome_modelo'] if variant.get('product_model') else "",
                        variant['product_model']['brands']['name'] if variant.get('product_model') and variant['product_model'].get('brands') else "",
                        variant['colors']['name'] if variant.get('colors') else "",
                        variant['sizes']['value'] if variant.get('sizes') else "",
                        stock
                    ),
                    tags=(variant['id'],)
                )

            self.lbl_warehouse_info.config(text=f"Armazém: {wh_name} - {len(response.data)} variação(ões)")
            self.lbl_summary.config(text=f"Total de itens em stock: {total_items}")

            self.app.db.audit(
                self.app.user,
                "VIEW_WAREHOUSE",
                "warehouse_stock",
                entity_pk=f"warehouse_id={wh_id}",
                details={"warehouse_name": wh_name}
            )

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar armazém.\n\n{e}")

    def copy_gtin(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atenção", "Seleciona um produto.")
            return

        gtin = self.tree.item(sel[0], "values")[0]
        self.tree.clipboard_clear()
        self.tree.clipboard_append(gtin)
        self.tree.update()
        messagebox.showinfo("OK", f"GTIN {gtin} copiado para clipboard!")

    def export_to_excel(self):
        if not self.tree.get_children():
            messagebox.showwarning("Atenção", "Não há resultados para exportar.")
            return

        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém primeiro.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"armazem_{wh_name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = wh_name[:31]

            headers = ["GTIN", "Modelo", "Marca", "Cor", "Tamanho", "Quantidade"]
            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            total_items = 0
            for item in self.tree.get_children():
                values = self.tree.item(item, "values")
                ws.append(values)
                total_items += int(values[5]) if values[5] else 0

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12

            ws.append([])
            ws.append(["Total de itens", total_items])
            resumo_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            resumo_font = Font(bold=True)
            for cell in ws[len(ws)]:
                cell.fill = resumo_fill
                cell.font = resumo_font

            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Relatório exportado para:\n{file_path}")

            self.app.db.audit(self.app.user, "EXPORT_WAREHOUSE_EXCEL", "warehouse_stock",
                            details={"warehouse": wh_name, "file": file_path, "rows": len(self.tree.get_children())})
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar.\n\n{e}")
