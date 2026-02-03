from __future__ import annotations

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import TYPE_CHECKING

from ui.components.helpers import safe_int

if TYPE_CHECKING:
    from ui.main import App


class BulkUpdateWindow(tk.Toplevel):
    """Janela para alteração em massa de stock"""

    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self.db = app.db
        self.title("Alteração em Massa de Stock")
        self.geometry("950x750")
        self.resizable(True, True)

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(main_frame, text="Alteração em Massa de Stock", font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 12))

        ttk.Label(main_frame, text="Adicionar Produtos:", font=("Segoe UI", 10, "bold")).grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(6, 4))

        ttk.Label(main_frame, text="Buscar por:").grid(row=2, column=0, sticky="w")
        self.var_search_type = tk.StringVar(value="ref_keyinvoice")

        search_frame = ttk.Frame(main_frame)
        search_frame.grid(row=2, column=1, columnspan=2, sticky="w")
        ttk.Radiobutton(search_frame, text="Ref KeyInvoice", variable=self.var_search_type,
                       value="ref_keyinvoice").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(search_frame, text="Ref WooCommerce", variable=self.var_search_type,
                       value="ref_woocommerce").pack(side="left")

        ttk.Label(main_frame, text="Código:").grid(row=3, column=0, sticky="w", pady=(8, 0))
        self.entry_code = ttk.Entry(main_frame, width=30)
        self.entry_code.grid(row=3, column=1, sticky="w", pady=(8, 0))
        self.entry_code.bind("<Return>", lambda e: self.add_product_by_code())

        btn_frame1 = ttk.Frame(main_frame)
        btn_frame1.grid(row=3, column=2, columnspan=2, sticky="w", pady=(8, 0), padx=(8, 0))
        ttk.Button(btn_frame1, text="Adicionar", command=self.add_product_by_code).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame1, text="Carregar Excel", command=self.load_from_excel).pack(side="left")
        ttk.Button(btn_frame1, text="Limpar Lista", command=self.clear_products).pack(side="left", padx=(8, 0))

        ttk.Label(main_frame, text="Produtos a Alterar:", font=("Segoe UI", 10, "bold")).grid(
            row=4, column=0, columnspan=4, sticky="w", pady=(8, 4))

        cols = ("gtin", "modelo", "marca", "cor", "tamanho", "stock")
        self.tree_products = ttk.Treeview(main_frame, columns=cols, show="headings", height=8)
        self.tree_products.grid(row=5, column=0, columnspan=4, sticky="nsew", pady=(0, 8))

        self.tree_products.heading("gtin", text="GTIN")
        self.tree_products.heading("modelo", text="Modelo")
        self.tree_products.heading("marca", text="Marca")
        self.tree_products.heading("cor", text="Cor")
        self.tree_products.heading("tamanho", text="Tamanho")
        self.tree_products.heading("stock", text="Stock Atual")

        self.tree_products.column("gtin", width=100, anchor="w")
        self.tree_products.column("modelo", width=120, anchor="w")
        self.tree_products.column("marca", width=80, anchor="w")
        self.tree_products.column("cor", width=70, anchor="w")
        self.tree_products.column("tamanho", width=60, anchor="w")
        self.tree_products.column("stock", width=150, anchor="w")

        tree_scroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.tree_products.yview)
        tree_scroll.grid(row=5, column=4, sticky="ns", pady=(0, 8))
        self.tree_products.config(yscrollcommand=tree_scroll.set)

        ttk.Separator(main_frame, orient="horizontal").grid(row=6, column=0, columnspan=5, sticky="ew", pady=8)

        ttk.Label(main_frame, text="Configuração da Operação:", font=("Segoe UI", 10, "bold")).grid(
            row=7, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Label(main_frame, text="Armazém:").grid(row=8, column=0, sticky="w", pady=4)
        self.var_warehouse = tk.StringVar()
        self.combo_warehouse = ttk.Combobox(main_frame, textvariable=self.var_warehouse, width=30, state="readonly")
        self.combo_warehouse.grid(row=8, column=1, columnspan=2, sticky="w", pady=4)

        ttk.Label(main_frame, text="Operação:").grid(row=9, column=0, sticky="w", pady=4)
        self.var_operation = tk.StringVar(value="add")

        op_frame = ttk.Frame(main_frame)
        op_frame.grid(row=9, column=1, columnspan=2, sticky="w", pady=4)
        ttk.Radiobutton(op_frame, text="Adicionar", variable=self.var_operation, value="add").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(op_frame, text="Remover", variable=self.var_operation, value="remove").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(op_frame, text="Definir", variable=self.var_operation, value="set").pack(side="left")

        ttk.Label(main_frame, text="Quantidade:").grid(row=10, column=0, sticky="w", pady=4)
        self.quantity = ttk.Entry(main_frame, width=15)
        self.quantity.grid(row=10, column=1, sticky="w", pady=4)

        ttk.Label(main_frame, text="(Para 'Adicionar' e 'Remover', usa a mesma quantidade para todos)",
                 font=("Segoe UI", 8), foreground="gray").grid(row=11, column=0, columnspan=3, sticky="w", pady=(0, 8))

        ttk.Separator(main_frame, orient="horizontal").grid(row=12, column=0, columnspan=5, sticky="ew", pady=8)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=13, column=0, columnspan=3, pady=10)

        ttk.Button(btn_frame, text="Processar", command=self.process_bulk_update, width=15).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy, width=15).pack(side="left", padx=5)

        ttk.Label(main_frame, text="Resultado:", font=("Segoe UI", 10, "bold")).grid(
            row=14, column=0, columnspan=3, sticky="w", pady=(8, 4))

        self.txt_log = tk.Text(main_frame, height=10, width=50, state="disabled", wrap="word")
        self.txt_log.grid(row=15, column=0, columnspan=3, sticky="nsew", pady=(0, 8))

        log_scroll = ttk.Scrollbar(main_frame, orient="vertical", command=self.txt_log.yview)
        log_scroll.grid(row=15, column=3, sticky="ns", pady=(0, 8))
        self.txt_log.config(yscrollcommand=log_scroll.set)

        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(5, weight=1)
        main_frame.rowconfigure(15, weight=1)

        self._load_warehouses()

    def _load_warehouses(self):
        try:
            response = self.db.supabase.table('warehouses').select('id, name').execute()
            warehouses = {row['name']: row['id'] for row in response.data}
            self.warehouse_name_to_id = warehouses
            self.combo_warehouse['values'] = list(warehouses.keys())
            if self.combo_warehouse['values']:
                self.var_warehouse.set(self.combo_warehouse['values'][0])
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar armazéns.\n\n{e}")

    def load_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            gtins = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    gtin = str(row[0]).strip()
                    if gtin:
                        gtins.append(gtin)

            found_count = 0
            not_found = []

            for gtin in gtins:
                success, msg = self._add_product_to_list(gtin, "gtin")
                if success:
                    found_count += 1
                else:
                    not_found.append(f"{gtin}: {msg}")

            msg = f"{found_count} produtos adicionados do Excel."
            if not_found:
                msg += f"\n\nNão encontrados:\n" + "\n".join(not_found[:10])
                if len(not_found) > 10:
                    msg += f"\n... e mais {len(not_found) - 10}"

            messagebox.showinfo("Carregamento Concluído", msg)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao carregar arquivo Excel.\n\n{e}")

    def add_product_by_code(self):
        code = self.entry_code.get().strip()
        if not code:
            messagebox.showwarning("Atenção", "Informa o código.")
            return

        search_type = self.var_search_type.get()
        success, msg = self._add_product_to_list(code, search_type)

        if success:
            self.entry_code.delete(0, tk.END)
            self.entry_code.focus()
        else:
            messagebox.showwarning("Não encontrado", msg)

    def _add_product_to_list(self, code, search_type):
        try:
            results = self.db.search_variants(code, search_type)
            if not results:
                return False, f"Código '{code}' não encontrado"

            added = 0
            for result in results:
                variant_id = result.get("variant_id")
                gtin = result.get("gtin") or ""
                modelo = result.get("nome_modelo") or "N/A"
                marca = result.get("marca") or "N/A"
                cor = result.get("cor") or "N/A"
                tamanho = result.get("tamanho") or "N/A"

                already_exists = False
                for item in self.tree_products.get_children():
                    if self.tree_products.item(item, "values")[0] == gtin:
                        already_exists = True
                        break

                if already_exists:
                    continue

                stock_response = self.db.supabase.table('warehouse_stock').select(
                    'stock, warehouses(name)'
                ).eq('variant_id', variant_id).execute()

                if stock_response.data:
                    stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                    stock_text = " | ".join(stocks)
                else:
                    stock_text = "Sem stock"

                self.tree_products.insert(
                    "", "end",
                    values=(gtin, modelo, marca, cor, tamanho, stock_text),
                    tags=(variant_id,)
                )
                added += 1

            if added == 0:
                return False, "Produto já está na lista"

            return True, f"{added} produto(s) adicionado(s)"
        except Exception as e:
            return False, f"Erro: {str(e)}"

    def clear_products(self):
        for item in self.tree_products.get_children():
            self.tree_products.delete(item)

    def log(self, message):
        self.txt_log.config(state="normal")
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state="disabled")
        self.update()

    def process_bulk_update(self):
        if not self.tree_products.get_children():
            messagebox.showwarning("Atenção", "Adiciona pelo menos um produto à lista.")
            return

        wh_name = self.var_warehouse.get()
        if not wh_name:
            messagebox.showwarning("Atenção", "Seleciona um armazém.")
            return

        operation = self.var_operation.get()
        quantity_str = self.quantity.get().strip()
        if not quantity_str:
            messagebox.showwarning("Atenção", "Informa a quantidade.")
            return

        quantity = safe_int(quantity_str, None)
        if quantity is None or quantity <= 0:
            messagebox.showwarning("Atenção", "Quantidade inválida.")
            return

        wh_id = self.warehouse_name_to_id.get(wh_name)
        products_count = len(self.tree_products.get_children())

        if not messagebox.askyesno(
            "Confirmar",
            f"Confirma alteração em massa?\n\n"
            f"Produtos: {products_count}\n"
            f"Armazém: {wh_name}\n"
            f"Operação: {operation}\n"
            f"Quantidade: {quantity}",
        ):
            return

        self.txt_log.config(state="normal")
        self.txt_log.delete("1.0", tk.END)
        self.txt_log.config(state="disabled")

        self.log("Iniciando alteração em massa...")
        self.log(f"Total de produtos: {products_count}\n")

        success_count = 0
        error_count = 0

        for item in self.tree_products.get_children():
            values = self.tree_products.item(item, "values")
            gtin = values[0]
            modelo = values[1]
            marca = values[2]
            cor = values[3]
            tamanho = values[4]
            variant_id = self.tree_products.item(item, "tags")[0]

            produto_info = f"{gtin} - {modelo} | {marca} | {cor} | Tam: {tamanho}"

            try:
                if operation == "add":
                    success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, quantity)
                elif operation == "remove":
                    success, msg = self.app.product_service.remove_from_stock(variant_id, wh_id, quantity)
                else:
                    current_response = self.db.supabase.table('warehouse_stock').select('stock').eq(
                        'variant_id', variant_id).eq('warehouse_id', wh_id).execute()
                    if current_response.data:
                        current_stock = current_response.data[0]['stock']
                        diff = quantity - current_stock
                        if diff > 0:
                            success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, diff)
                        elif diff < 0:
                            success, msg = self.app.product_service.remove_from_stock(variant_id, wh_id, abs(diff))
                        else:
                            success, msg = True, f"Stock já está em {quantity}"
                    else:
                        success, msg = self.app.product_service.add_to_stock(variant_id, wh_id, quantity)

                if success:
                    self.log(f"✓ {produto_info}")
                    self.log(f"  → {msg}\n")
                    success_count += 1
                    self.app.db.audit(
                        self.app.user,
                        f"BULK_{operation.upper()}_STOCK",
                        "warehouse_stock",
                        entity_pk=f"variant_id={variant_id},warehouse_id={wh_id}",
                        details={"gtin": gtin, "quantity": quantity, "operation": operation},
                    )
                else:
                    self.log(f"❌ {produto_info}")
                    self.log(f"  → {msg}\n")
                    error_count += 1
            except Exception as e:
                self.log(f"❌ GTIN {gtin}: Erro - {str(e)}\n")
                error_count += 1

        self.log(f"\n{'='*50}")
        self.log("Processamento concluído!")
        self.log(f"Sucesso: {success_count}")
        self.log(f"Erros: {error_count}")
        self.log(f"Total: {products_count}")

        messagebox.showinfo(
            "Concluído",
            "Alteração em massa concluída!\n\n"
            f"Sucesso: {success_count}\n"
            f"Erros: {error_count}\n"
            f"Total: {products_count}",
        )
