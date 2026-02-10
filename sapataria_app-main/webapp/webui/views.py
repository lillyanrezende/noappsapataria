from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse

from db import DB
from services import ProductService, AuthService, DomainService


def safe_int(value, default=None):
    try:
        return int(str(value).strip())
    except Exception:
        return default


def get_services():
    db = DB()
    return db, ProductService(db), AuthService(db), DomainService(db)


def get_domains(domain_service):
    return {
        "brands": domain_service.get_domain_list("brands"),
        "categories": domain_service.get_domain_list("categories"),
        "colors": domain_service.get_domain_list("colors"),
        "sizes": domain_service.get_domain_list("sizes"),
        "warehouses": domain_service.get_domain_list("warehouses"),
        "suppliers": domain_service.get_domain_list("suppliers"),
    }


def require_login(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get("user"):
            return redirect("webui:login")
        return view_func(request, *args, **kwargs)

    return wrapper


def login_view(request):
    _, _, auth_service, _ = get_services()
    context = {"error": None}

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        if not username or not password:
            context["error"] = "Username and password are required."
        else:
            try:
                res = auth_service.authenticate(username, password)
                if res == "inactive":
                    context["error"] = "User is inactive."
                elif not res:
                    context["error"] = "Invalid credentials."
                else:
                    request.session["user"] = res
                    return redirect("webui:create")
            except Exception as exc:
                context["error"] = f"Login failed: {exc}"

    return render(request, "webui/login.html", context)


def logout_view(request):
    user = request.session.get("user")
    if user:
        db, _, _, _ = get_services()
        db.audit(user, "LOGOUT", "profiles", entity_pk=f"user_id={user['user_id']}", details={})
    request.session.flush()
    return redirect("webui:login")


def register_view(request):
    _, _, auth_service, _ = get_services()
    context = {"error": None, "success": None}

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        nome_usuario = request.POST.get("nome_usuario", "").strip()
        setor = request.POST.get("setor", "").strip()
        password = request.POST.get("password", "").strip()

        if not username or not nome_usuario or not password:
            context["error"] = "Username, name, and password are required."
        else:
            try:
                user_id = auth_service.create_user(username, nome_usuario, setor, password)
                db, _, _, _ = get_services()
                db.audit(
                    {"user_id": user_id, "username": username},
                    "REGISTER",
                    "profiles",
                    entity_pk=f"user_id={user_id}",
                    details={"username": username, "setor": setor},
                )
                context["success"] = "Account created. You can login now."
            except Exception as exc:
                context["error"] = f"Register failed: {exc}"

    return render(request, "webui/register.html", context)


@require_login
def create_view(request):
    db, product_service, _, domain_service = get_services()
    user = request.session.get("user")
    messages = []
    form = request.POST if request.method == "POST" else {}

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add_domain":
            table = request.POST.get("domain_table")
            value = request.POST.get("domain_value", "").strip()
            category_id = request.POST.get("domain_category_id")
            if not table or not value:
                messages.append(("error", "Provide domain table and value."))
            else:
                try:
                    if table == "subcategories":
                        if not category_id:
                            messages.append(("error", "Category is required for subcategory."))
                        else:
                            domain_service.add_subcategory(category_id, value)
                            messages.append(("success", "Subcategory added."))
                    else:
                        domain_service.add_domain_value(table, value)
                        messages.append(("success", "Domain value added."))
                except Exception as exc:
                    messages.append(("error", f"Domain add failed: {exc}"))

        if action == "create_product":
            gtin = request.POST.get("gtin", "").strip()
            nome_modelo = request.POST.get("nome_modelo", "").strip()
            stock_val = safe_int(request.POST.get("stock"), None)

            if not gtin:
                messages.append(("error", "GTIN is required."))
            elif not nome_modelo:
                messages.append(("error", "Model name is required."))
            elif stock_val is None or stock_val < 0:
                messages.append(("error", "Invalid stock value."))
            else:
                try:
                    success, message, variant_id = product_service.create_or_update_product(
                        gtin=gtin,
                        nome_modelo=nome_modelo,
                        brand_id=request.POST.get("brand_id"),
                        category_id=request.POST.get("category_id"),
                        subcategory_id=request.POST.get("subcategory_id"),
                        supplier_id=request.POST.get("supplier_id"),
                        color_id=request.POST.get("color_id"),
                        size_id=request.POST.get("size_id"),
                        warehouse_id=request.POST.get("warehouse_id"),
                        stock=stock_val,
                        ref_keyinvoice=request.POST.get("ref_keyinvoice") or None,
                        ref_woocommerce=request.POST.get("ref_woocommerce") or None,
                    )
                    if success:
                        db.audit(
                            user,
                            "CREATE_OR_UPDATE_PRODUCT",
                            "product_variant",
                            entity_pk=f"id={variant_id}",
                            details={"gtin": gtin, "stock": stock_val},
                        )
                        messages.append(("success", message))
                    else:
                        messages.append(("error", message))
                except Exception as exc:
                    messages.append(("error", f"Create failed: {exc}"))

    domains = get_domains(domain_service)

    category_id = (
        request.POST.get("category_id")
        or request.GET.get("category_id")
        or (domains["categories"][0][0] if domains["categories"] else None)
    )
    subcategories = domain_service.get_subcategories_by_category(category_id) if category_id else []

    context = {
        "user": user,
        "messages": messages,
        "form": form,
        "domains": domains,
        "subcategories": subcategories,
    }
    return render(request, "webui/create.html", context)


@require_login
def update_view(request):
    db, product_service, _, domain_service = get_services()
    user = request.session.get("user")
    messages = []
    search_results = []
    loaded = None
    stocks = []
    form = request.POST if request.method == "POST" else {}

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "search":
            value = request.POST.get("search_value", "").strip()
            search_type = request.POST.get("search_type", "gtin")
            if not value:
                messages.append(("error", "Provide a search value."))
            else:
                try:
                    search_results = db.search_variants(value, search_type)
                    for item in search_results:
                        stock_response = db.supabase.table("warehouse_stock").select(
                            "stock, warehouses(name)"
                        ).eq("variant_id", item["variant_id"]).execute()
                        if stock_response.data:
                            stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                            item["stock"] = " | ".join(stocks)
                        else:
                            item["stock"] = "Sem stock"
                    if not search_results:
                        messages.append(("info", "No variants found."))
                except Exception as exc:
                    messages.append(("error", f"Search failed: {exc}"))

        if action == "load_variant":
            variant_id = request.POST.get("variant_id")
            if variant_id:
                full = db.get_full_view_by_variant_id(variant_id)
                if full:
                    loaded, stocks = full
                else:
                    messages.append(("info", "Variant not found."))

        if action == "save_item":
            gtin = request.POST.get("gtin", "").strip()
            if not gtin:
                messages.append(("error", "Load a variant first."))
            else:
                try:
                    ok, msg = product_service.update_product_details(
                        gtin=gtin,
                        model_fields={
                            "nome_modelo": request.POST.get("nome_modelo", ""),
                            "marca_id": request.POST.get("brand_id"),
                            "categoria_id": request.POST.get("category_id"),
                            "subcategoria_id": request.POST.get("subcategory_id"),
                            "fornecedor_id": request.POST.get("supplier_id"),
                        },
                        variant_fields={
                            "cor_id": request.POST.get("color_id"),
                            "tamanho_id": request.POST.get("size_id"),
                            "ref_keyinvoice": request.POST.get("ref_keyinvoice") or None,
                        },
                    )
                    if ok:
                        db.audit(
                            user,
                            "UPDATE_ITEM",
                            "product_variant",
                            entity_pk=f"gtin={gtin}",
                            details={"fields": ["model", "variant"]},
                        )
                        messages.append(("success", msg))
                    else:
                        messages.append(("error", msg))
                except Exception as exc:
                    messages.append(("error", f"Update failed: {exc}"))

        if action in ("add_stock", "remove_stock"):
            variant_id = request.POST.get("variant_id")
            warehouse_id = request.POST.get("warehouse_id")
            quantity = safe_int(request.POST.get("stock"), None)
            if not variant_id:
                messages.append(("error", "Load a variant first."))
            elif quantity is None or quantity <= 0:
                messages.append(("error", "Invalid quantity."))
            elif not warehouse_id:
                messages.append(("error", "Select a warehouse."))
            else:
                try:
                    if action == "add_stock":
                        success, msg = product_service.add_to_stock(variant_id, warehouse_id, quantity)
                    else:
                        success, msg = product_service.remove_from_stock(variant_id, warehouse_id, quantity)
                    if success:
                        db.audit(
                            user,
                            "ADD_STOCK" if action == "add_stock" else "REMOVE_STOCK",
                            "warehouse_stock",
                            entity_pk=f"variant_id={variant_id},warehouse_id={warehouse_id}",
                            details={"quantity": quantity},
                        )
                        messages.append(("success", msg))
                    else:
                        messages.append(("error", msg))
                except Exception as exc:
                    messages.append(("error", f"Stock update failed: {exc}"))

        if action in ("add_stock", "remove_stock", "save_item"):
            variant_id = request.POST.get("variant_id")
            if variant_id:
                full = db.get_full_view_by_variant_id(variant_id)
                if full:
                    loaded, stocks = full

    domains = get_domains(domain_service)

    category_id = (
        request.POST.get("category_id")
        or (loaded["categoria_id"] if loaded else None)
        or (domains["categories"][0][0] if domains["categories"] else None)
    )
    subcategories = domain_service.get_subcategories_by_category(category_id) if category_id else []

    context = {
        "user": user,
        "messages": messages,
        "form": form,
        "domains": domains,
        "subcategories": subcategories,
        "search_results": search_results,
        "loaded": loaded,
        "stocks": stocks,
    }
    return render(request, "webui/update.html", context)


@require_login
def delete_view(request):
    db, product_service, _, domain_service = get_services()
    user = request.session.get("user")
    messages = []
    loaded = None
    stocks = []
    form = request.POST if request.method == "POST" else {}

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "preview":
            value = request.POST.get("search_value", "").strip()
            search_type = request.POST.get("search_type", "gtin")
            if not value:
                messages.append(("error", "Provide a code."))
            else:
                try:
                    full = db.get_full_view_by_gtin(value, search_type)
                    if full:
                        loaded, stocks = full
                    else:
                        messages.append(("info", "Code not found."))
                except Exception as exc:
                    messages.append(("error", f"Preview failed: {exc}"))

        if action == "delete_stock":
            variant_id = request.POST.get("variant_id")
            warehouse_id = request.POST.get("warehouse_id")
            if not variant_id or not warehouse_id:
                messages.append(("error", "Load a product and choose a warehouse."))
            else:
                try:
                    success, msg, deleted_variant = product_service.delete_stock(variant_id, warehouse_id)
                    db.audit(
                        user,
                        "DELETE_STOCK",
                        "warehouse_stock",
                        entity_pk=f"variant_id={variant_id},warehouse_id={warehouse_id}",
                        details={},
                    )
                    if deleted_variant:
                        db.audit(
                            user,
                            "DELETE_VARIANT",
                            "product_variant",
                            entity_pk=f"id={variant_id}",
                            details={"reason": "no_stock_rows"},
                        )
                    messages.append(("success" if success else "error", msg))
                except Exception as exc:
                    messages.append(("error", f"Delete failed: {exc}"))

    domains = get_domains(domain_service)

    context = {
        "user": user,
        "messages": messages,
        "form": form,
        "domains": domains,
        "loaded": loaded,
        "stocks": stocks,
    }
    return render(request, "webui/delete.html", context)


@require_login
def view_view(request):
    db, _, _, _ = get_services()
    user = request.session.get("user")
    messages = []
    results = []
    details = None
    details_stocks = []
    form = request.POST if request.method == "POST" else {}

    if request.method == "POST":
        action = request.POST.get("action")
        search_value = request.POST.get("search_value", "").strip()
        search_type = request.POST.get("search_type", "gtin")

        if action == "search":
            if not search_value:
                messages.append(("error", "Provide a search value."))
            else:
                try:
                    results = db.search_variants(search_value, search_type)
                    for item in results:
                        full = db.get_full_view_by_variant_id(item["variant_id"])
                        if full:
                            header, _ = full
                            item["marca"] = header.get("marca")
                        stock_response = db.supabase.table("warehouse_stock").select(
                            "stock, warehouses(name)"
                        ).eq("variant_id", item["variant_id"]).execute()
                        if stock_response.data:
                            stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                            item["stock"] = " | ".join(stocks)
                        else:
                            item["stock"] = "Sem stock"
                    if not results:
                        messages.append(("info", "No results."))
                except Exception as exc:
                    messages.append(("error", f"Search failed: {exc}"))

        if action == "details":
            variant_id = request.POST.get("variant_id")
            if variant_id:
                full = db.get_full_view_by_variant_id(variant_id)
                if full:
                    details, details_stocks = full
                    db.audit(
                        user,
                        "VIEW_DETAILS",
                        "product_variant",
                        entity_pk=f"id={variant_id}",
                        details={},
                    )
                else:
                    messages.append(("info", "Variant not found."))

        if action == "export":
            if not search_value:
                messages.append(("error", "Provide a search value to export."))
            else:
                try:
                    results = db.search_variants(search_value, search_type)
                    for item in results:
                        full = db.get_full_view_by_variant_id(item["variant_id"])
                        if full:
                            header, _ = full
                            item["marca"] = header.get("marca")
                    wb_stream = build_excel_for_variants(db, results)
                    response = HttpResponse(
                        wb_stream.getvalue(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    response["Content-Disposition"] = "attachment; filename=relatorio_produtos.xlsx"
                    return response
                except Exception as exc:
                    messages.append(("error", f"Export failed: {exc}"))

    context = {
        "user": user,
        "messages": messages,
        "form": form,
        "results": results,
        "details": details,
        "details_stocks": details_stocks,
    }
    return render(request, "webui/view.html", context)


@require_login
def warehouse_view(request):
    db, _, _, domain_service = get_services()
    user = request.session.get("user")
    messages = []
    rows = []
    summary = None
    form = request.POST if request.method == "POST" else {}

    warehouses = domain_service.get_domain_list("warehouses")

    if request.method == "POST":
        action = request.POST.get("action")
        warehouse_id = request.POST.get("warehouse_id")

        if action in ("load", "export"):
            if not warehouse_id:
                messages.append(("error", "Select a warehouse."))
            else:
                try:
                    warehouse_name = None
                    for item in warehouses:
                        if str(item[0]) == str(warehouse_id):
                            warehouse_name = item[1]
                            break
                    rows = []
                    page_size = 1000
                    offset = 0
                    while True:
                        response = db.supabase.table("warehouse_stock").select(
                            "stock, product_variant(id, gtin, product_model(nome_modelo, brands(name), categories(name), subcategories(name)), colors(name), sizes(value))"
                        ).eq("warehouse_id", warehouse_id).order("product_variant(gtin)").range(
                            offset, offset + page_size - 1
                        ).execute()

                        batch = response.data or []
                        if not batch:
                            break
                        rows.extend(batch)
                        if len(batch) < page_size:
                            break
                        offset += page_size
                    total_items = sum([row.get("stock", 0) for row in rows])
                    summary = {"total_items": total_items, "count": len(rows)}

                    if action == "export":
                        wb_stream = build_excel_for_warehouse(rows)
                        response = HttpResponse(
                            wb_stream.getvalue(),
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        safe_name = (warehouse_name or "armazem").lower().replace(" ", "_")
                        response["Content-Disposition"] = f"attachment; filename=armazem_{safe_name}.xlsx"
                        db.audit(
                            user,
                            "EXPORT_WAREHOUSE_EXCEL",
                            "warehouse_stock",
                            details={"warehouse_id": warehouse_id, "rows": len(rows)},
                        )
                        return response
                except Exception as exc:
                    messages.append(("error", f"Load failed: {exc}"))

    context = {
        "user": user,
        "messages": messages,
        "form": form,
        "warehouses": warehouses,
        "rows": rows,
        "summary": summary,
    }
    return render(request, "webui/warehouse.html", context)


@require_login
def bulk_update_view(request):
    db, product_service, _, domain_service = get_services()
    user = request.session.get("user")
    messages = []
    logs = []
    codes_text = request.POST.get("codes", "") if request.method == "POST" else ""
    form = request.POST if request.method == "POST" else {}
    preview_rows = []
    details = None
    details_stocks = []

    warehouses = domain_service.get_domain_list("warehouses")

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "preview_codes":
            search_type = request.POST.get("search_type", "ref_keyinvoice")
            codes = []
            single_code = request.POST.get("single_code", "").strip()
            if single_code:
                codes.append(single_code)
            raw_codes = request.POST.get("codes", "")
            if raw_codes:
                codes.extend(parse_codes_from_text(raw_codes))
            if not codes:
                messages.append(("error", "Provide codes to search."))
            else:
                preview_rows = build_bulk_preview(db, codes, search_type)
                codes_text = build_bulk_codes_text(preview_rows)

        if action == "preview_excel":
            if not request.FILES.get("excel"):
                messages.append(("error", "Select an Excel file to load."))
            else:
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(request.FILES["excel"], read_only=True)
                    ws = wb.active
                    codes = []
                    headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                    search_type = request.POST.get("search_type", "ref_keyinvoice")
                    code_index = find_excel_code_index(headers, search_type)
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > code_index and row[code_index]:
                            code = str(row[code_index]).strip()
                            if code:
                                codes.append(code)
                    preview_rows = build_bulk_preview(db, codes, search_type)
                    codes_text = build_bulk_codes_text(preview_rows)
                    messages.append(("success", f"Carregados {len(codes)} codigos do Excel."))
                except Exception as exc:
                    messages.append(("error", f"Excel load failed: {exc}"))

        if action == "details":
            variant_id = request.POST.get("variant_id")
            if variant_id:
                full = db.get_full_view_by_variant_id(variant_id)
                if full:
                    details, details_stocks = full
                else:
                    messages.append(("info", "Variant not found."))
            search_type = request.POST.get("search_type", "ref_keyinvoice")
            if codes_text:
                codes = parse_codes_from_text(codes_text)
                preview_rows = build_bulk_preview(db, codes, search_type)
                codes_text = build_bulk_codes_text(preview_rows)

        if action == "process":
            warehouse_id = request.POST.get("warehouse_id")
            operation = request.POST.get("operation", "add")
            quantity = safe_int(request.POST.get("quantity"), None)
            search_type = request.POST.get("search_type", "ref_keyinvoice")

            if not warehouse_id:
                messages.append(("error", "Select a warehouse."))
            elif quantity is None or quantity <= 0:
                messages.append(("error", "Provide a valid quantity."))
            else:
                codes = []
                raw_codes = request.POST.get("codes", "")
                if raw_codes:
                    codes.extend(parse_codes_from_text(raw_codes))

                if request.FILES.get("excel"):
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(request.FILES["excel"], read_only=True)
                        ws = wb.active
                        headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                        code_index = find_excel_code_index(headers, search_type)
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and len(row) > code_index and row[code_index]:
                                code = str(row[code_index]).strip()
                                if code:
                                    codes.append(code)
                    except Exception as exc:
                        messages.append(("error", f"Excel load failed: {exc}"))

                if not codes:
                    messages.append(("error", "Provide codes or an Excel file."))
                else:
                    if search_type != "gtin" and all(code.isdigit() and len(code) in (13, 14) for code in codes):
                        search_type = "gtin"
                    success_count = 0
                    error_count = 0

                    for code in codes:
                        try:
                            results = db.search_variants(code, search_type)
                            if not results:
                                logs.append(f"Not found: {code}")
                                error_count += 1
                                continue

                            for result in results:
                                variant_id = result.get("variant_id")
                                if operation == "add":
                                    success, msg = product_service.add_to_stock(variant_id, warehouse_id, quantity)
                                elif operation == "remove":
                                    success, msg = product_service.remove_from_stock(variant_id, warehouse_id, quantity)
                                else:
                                    current = db.supabase.table("warehouse_stock").select("stock").eq(
                                        "variant_id", variant_id
                                    ).eq("warehouse_id", warehouse_id).execute()
                                    if current.data:
                                        current_stock = current.data[0]["stock"]
                                        diff = quantity - current_stock
                                        if diff > 0:
                                            success, msg = product_service.add_to_stock(variant_id, warehouse_id, diff)
                                        elif diff < 0:
                                            success, msg = product_service.remove_from_stock(variant_id, warehouse_id, abs(diff))
                                        else:
                                            success, msg = True, f"Stock already at {quantity}"
                                    else:
                                        success, msg = product_service.add_to_stock(variant_id, warehouse_id, quantity)

                                if success:
                                    logs.append(f"OK: {code} -> {msg}")
                                    success_count += 1
                                else:
                                    logs.append(f"ERROR: {code} -> {msg}")
                                    error_count += 1

                                db.audit(
                                    user,
                                    f"BULK_{operation.upper()}_STOCK",
                                    "warehouse_stock",
                                    entity_pk=f"variant_id={variant_id},warehouse_id={warehouse_id}",
                                    details={"code": code, "quantity": quantity, "operation": operation},
                                )
                        except Exception as exc:
                            logs.append(f"ERROR: {code} -> {exc}")
                            error_count += 1

                    messages.append(("success", f"Processed. Success: {success_count}. Errors: {error_count}."))

        if not preview_rows and codes_text:
            search_type = request.POST.get("search_type", "ref_keyinvoice")
            codes = parse_codes_from_text(codes_text)
            if codes:
                preview_rows = build_bulk_preview(db, codes, search_type)
                codes_text = build_bulk_codes_text(preview_rows)

    context = {
        "user": user,
        "messages": messages,
        "logs": logs,
        "warehouses": warehouses,
        "codes_text": codes_text,
        "form": form,
        "preview_rows": preview_rows,
        "details": details,
        "details_stocks": details_stocks,
    }
    return render(request, "webui/bulk_update.html", context)


def build_bulk_preview(db, codes, search_type):
    rows = []
    for code in codes:
        results = db.search_variants(code, search_type)
        if not results:
            continue
        for result in results:
            variant_id = result.get("variant_id")
            full = db.get_full_view_by_variant_id(variant_id)
            brand = None
            category = None
            subcategory = None
            if full:
                header, _ = full
                brand = header.get("marca")
                category = header.get("categoria")
                subcategory = header.get("subcategoria")

            stock_response = db.supabase.table("warehouse_stock").select(
                "stock, warehouses(name)"
            ).eq("variant_id", variant_id).execute()

            stock_rows = []
            if stock_response.data:
                stocks = [f"{s['warehouses']['name']}: {s['stock']}" for s in stock_response.data]
                stock_text = " | ".join(stocks)
                stock_rows = [
                    {"warehouse": s["warehouses"]["name"], "stock": s["stock"]}
                    for s in stock_response.data
                ]
            else:
                stock_text = "Sem stock"

            rows.append(
                {
                    "variant_id": variant_id,
                    "gtin": result.get("gtin"),
                    "category": category,
                    "subcategory": subcategory,
                    "nome_modelo": result.get("nome_modelo"),
                    "marca": brand,
                    "cor": result.get("cor"),
                    "tamanho": result.get("tamanho"),
                    "ref_keyinvoice": result.get("ref_keyinvoice"),
                    "ref_woocomerce": result.get("ref_woocomerce"),
                    "stock": stock_text,
                    "stock_rows": stock_rows,
                }
            )
    return rows


def build_bulk_codes_text(preview_rows):
    header = [
        "GTIN",
        "Ref Keyinvoice",
        "Ref Woocomerce",
        "Categoria",
        "Subcategoria",
        "Modelo",
        "Marca",
        "Cor",
        "Tamanho",
    ]

    warehouse_names = []
    for row in preview_rows:
        for stock in row.get("stock_rows") or []:
            name = stock.get("warehouse")
            if name and name not in warehouse_names:
                warehouse_names.append(name)

    header.extend([f"Quantidade Armazem {name}" for name in warehouse_names])
    lines = [";".join(header)]

    for row in preview_rows:
        stock_map = {}
        for stock in row.get("stock_rows") or []:
            stock_map[stock.get("warehouse")] = stock.get("stock")

        line = [
            str(row.get("gtin") or ""),
            str(row.get("ref_keyinvoice") or ""),
            str(row.get("ref_woocomerce") or ""),
            str(row.get("category") or ""),
            str(row.get("subcategory") or ""),
            str(row.get("nome_modelo") or ""),
            str(row.get("marca") or ""),
            str(row.get("cor") or ""),
            str(row.get("tamanho") or ""),
        ]
        line.extend([str(stock_map.get(name, 0)) for name in warehouse_names])
        lines.append(";".join(line))

    return "\n".join(lines)


def parse_codes_from_text(text):
    codes = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(";")]
        if not parts:
            continue
        if parts[0].lower() == "gtin":
            continue
        if parts[0]:
            codes.append(parts[0])
    return codes


def find_excel_code_index(headers, search_type):
    normalized = [h.strip().lower() for h in headers]
    if search_type == "gtin":
        candidates = ["gtin"]
    elif search_type == "ref_keyinvoice":
        candidates = ["ref keyinvoice", "ref key invoice", "refkeyinvoice"]
    else:
        candidates = ["ref woocommerce", "ref woocomerce", "ref_woocommerce", "ref_woocomerce"]

    for candidate in candidates:
        if candidate in normalized:
            return normalized.index(candidate)

    # fallback to first column
    return 0


def subcategories_view(request):
    _, _, _, domain_service = get_services()
    category_id = request.GET.get("category_id")
    if not category_id:
        return JsonResponse({"items": []})
    items = domain_service.get_subcategories_by_category(category_id)
    data = [{"id": row[0], "name": row[1]} for row in items]
    return JsonResponse({"items": data})


def build_excel_for_variants(db, results):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    warehouses_resp = db.supabase.table("warehouses").select("id, name").order("name").execute()
    warehouses = warehouses_resp.data or []
    warehouse_names = [w["name"] for w in warehouses]

    headers = [
        "GTIN",
        "Ref KeyInvoice",
        "Ref WooCommerce",
        "Categoria",
        "Subcategoria",
        "Modelo",
        "Marca",
        "Cor",
        "Tamanho",
    ]
    headers.extend([f"Quantidade Armazem {name}" for name in warehouse_names])
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        full = db.get_full_view_by_variant_id(r["variant_id"])
        header = full[0] if full else {}

        stock_response = db.supabase.table("warehouse_stock").select(
            "stock, warehouses(name)"
        ).eq("variant_id", r["variant_id"]).execute()

        stock_map = {}
        for s in (stock_response.data or []):
            stock_map[s["warehouses"]["name"]] = s["stock"]

        row = [
            r.get("gtin") or "",
            r.get("ref_keyinvoice") or "",
            r.get("ref_woocomerce") or "",
            header.get("categoria") or "",
            header.get("subcategoria") or "",
            r.get("nome_modelo") or "",
            header.get("marca") or r.get("marca") or "",
            r.get("cor") or "",
            r.get("tamanho") or "",
        ]
        row.extend([stock_map.get(name, 0) for name in warehouse_names])
        ws.append(row)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_excel_for_warehouse(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Armazem"

    headers = ["GTIN", "Modelo", "Marca", "Cor", "Tamanho", "Quantidade"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        variant = row.get("product_variant") or {}
        model = variant.get("product_model") or {}
        brand = model.get("brands") or {}
        color = variant.get("colors") or {}
        size = variant.get("sizes") or {}

        ws.append(
            [
                variant.get("gtin") or "",
                model.get("nome_modelo") or "",
                brand.get("name") or "",
                color.get("name") or "",
                size.get("value") or "",
                row.get("stock") or 0,
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
