"""Script para criar arquivo Excel de exemplo para alteração em massa"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "GTINs para Atualizar"

# Header
headers = ["GTIN", "Modelo (referência)", "Observação"]
ws.append(headers)

# Estilo do header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Dados de exemplo (você pode substituir pelos GTINs reais)
example_data = [
    ["1234567890123", "Modelo A", "Exemplo 1"],
    ["9876543210987", "Modelo B", "Exemplo 2"],
    ["5555555555555", "Modelo C", "Exemplo 3"],
    ["1111111111111", "Modelo D", "Exemplo 4"],
    ["2222222222222", "Modelo E", "Exemplo 5"],
]

for row_data in example_data:
    ws.append(row_data)

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 30

# Adicionar instruções
ws2 = wb.create_sheet("Instruções")
instructions = [
    ["Instruções para uso do arquivo de alteração em massa:"],
    [""],
    ["1. Na aba 'GTINs para Atualizar', preencha a coluna A com os códigos GTIN dos produtos"],
    ["2. As colunas B e C são apenas para referência, não são usadas pelo sistema"],
    ["3. A primeira linha (header) será ignorada"],
    ["4. No programa, clique em 'Alteração em Massa' na aba 'Alterar'"],
    ["5. Clique em 'Carregar de Excel' e selecione este arquivo"],
    ["6. Escolha o armazém, operação (Adicionar/Remover/Definir) e quantidade"],
    ["7. Clique em 'Processar' para executar"],
    [""],
    ["Tipos de operação:"],
    ["- Adicionar: Soma a quantidade ao stock existente"],
    ["- Remover: Subtrai a quantidade do stock existente"],
    ["- Definir: Define o stock para o valor especificado (independente do valor atual)"],
]

for instruction in instructions:
    ws2.append(instruction)

ws2.column_dimensions['A'].width = 80

# Salvar
wb.save("exemplo_bulk_update.xlsx")
print("Arquivo 'exemplo_bulk_update.xlsx' criado com sucesso!")
print("Use este arquivo como template para alteração em massa de stock.")
